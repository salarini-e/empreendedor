from django.shortcuts import render, get_object_or_404
from django.db import connection
from .api import ApiProtocolo
from .views_folder.minha_empresa import *
from .views_folder.vitrine_virtual import *
from .views_folder.admin import *
from .forms import Faccao_Legal_Form, Escola_Form, Solicitacao_de_Compras_Form,Criar_Item_Solicitacao, Criar_Processo_Docs_Form, RequerimentoISSQNForm, DocumentosPedidoForm, Processo_ISS_Form, Contrato_NotaFiscal, Contrato_Avaliacao, Form_Novas_Oportunidades, Form_Credito_Facil, Form_Necessidades_das_Empresas, FormRamos, Form_Natal_Artesao
from django.urls import reverse
from autenticacao.functions import validate_cpf
from .models import Profissao, Escola, Solicitacao_de_Compras, Item_Solicitacao, Proposta, Proposta_Item, Contrato_de_Servico, Tipo_Processos, Processo_Status_Documentos_Anexos, RequerimentoISS, AtividadeManual, Tipo_Producao_Alimentos, Tipo_Costura, Tipo_Producao_Bebidas, Faccao_legal, Novas_Oportunidades
from .functions.pdde import Listar_Proposta, PDDE_POST
from .functions.email import send_email_for_create_process, send_email_for_att_process
from .functions.empresa import validate_CNPJ
from django.db import transaction
# from guardiao.models import TentativaBurla, ErroPrevisto
from datetime import datetime
from django.utils import timezone

from django.http import HttpResponse
from openpyxl import Workbook
from .models import Empresa
# import pandas as pd
from autenticacao.models import User

# Create your views here.
def index(request):
    context = {
         'titulo': 'Sala do Empreendedor - Página Inicial',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/index.html', context)

def cursos(request):
    context = {
         'titulo': 'Sala do Empreendedor - Capacitação',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/cursos.html', context)


def conheca_nossa_sala(request):
    context = { 
        'titulo': 'Sala do Empreendedor - Conheça nossa sala',   
    }
    return render(request, 'sala_do_empreendedor/conheca-nossa-sala.html', context)

def em_construcao(request):
    context = {
         'titulo': 'Sala do Empreendedor - Em construção',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/em-construcao.html', context)

def consultar_protocolo(request):
    api = ApiProtocolo()
    status, response = api.recuperarAssuntos()
    print(response)
    message=response['message']
    if request.method == 'POST':
        
        status, message = api.recuperarProcesso(request.POST['cpf'])
        
    context = {
         'titulo': 'Sala do Empreendedor - Consultar Protocolo',
         'status': status,
         'message': message
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/consultar-protocolo.html', context)


def faccao_legal(request):
    context = {
        'titulo': 'Sala do Empreendedor - Facção Legal',
        'titulo_pag':'Facção Legal',
    }
    if request.user.is_authenticated:
        try:            
            faccao=Faccao_legal.objects.get(user=request.user)
            context['faccao']=faccao              
        except Exception as E:
            print(E)
    if request.method == 'POST':
        try:
            pessoa=Pessoa.objects.get(cpf=validate_cpf(request.POST['cpf']))
            messages.warning(request, 'Faça seu login antes de cadastrar sua facção')
            next_page = reverse('autenticacao:login') + f'?next={reverse("empreendedor:cadastrar_faccao_legal")}'
            return redirect(next_page)
        except Exception as E:
            print(E)
            next_page = reverse('autenticacao:cadastrar_usuario') + f'?next={reverse("empreendedor:cadastrar_faccao_legal")}'
            return redirect(next_page)

    return render(request, 'sala_do_empreendedor/faccao_legal.html', context)

@login_required
def apagar_faccao(request):
    try:
        faccao=Faccao_legal.objects.get(user=request.user)
        faccao.delete()
        messages.success(request, 'Facção apagada com sucesso!')
    except Exception as E:
        # ErroPrevisto.objects.create(
        #     local_deteccao='empreendedor:apagar_faccao',
        #     user=request.user,
        #     ip_address=request.META.get('REMOTE_ADDR'),
        #     informacoes_adicionais=f'Tentativa de apagar facção sem ter uma cadastrada ou com mais de um registro. Exception: {E}'
        # )
        messages.error(request, 'Erro ao apagar facção.')
    return redirect('empreendedor:faccao_legal')

@login_required
def export_faccoes(request):
    if request.user.is_superuser:
        response = HttpResponse(content_type='application/ms-excel')
        data = datetime.now().strftime('%d-%m-%Y')
        response['Content-Disposition'] = f'attachment; filename="faccoes-{data}.xls"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Facções"
        ws.append(['id_cadastro','Usuário', 'Telefone', 'Email', 'Possui MEI?', 'Tempo que trabalha com facção?', 
                'Trabalha com', 'Equipamentos', 'Área de trabalho separada?', 'Tamanho área de trabalho', 'Qnt. colaboradores',
                'Tipo de produto', ' Outro produto', 'Como está de trabalho?', 'Como considera a renda?',
                'Preferência', 'Sonho no setor'])
        faccoes = Faccao_legal.objects.all()    
        for faccao in faccoes:
            pessoa = Pessoa.objects.get(user=faccao.user)
            trabalha_com = ', '.join([str(item) for item in faccao.trabalha_com.all()])
            tipo_produto = ', '.join([str(item) for item in faccao.tipo_produto.all()])
            equipamentos = ', '.join([str(item) for item in faccao.equipamentos.all()])
            ws.append([faccao.id, pessoa.nome, pessoa.telefone, pessoa.email, faccao.possui_mei, faccao.get_tempo_que_trabalha_display(), trabalha_com, equipamentos, faccao.get_area_display(), faccao.get_tamanho_area_display(), faccao.qtd_colaboradores, tipo_produto, faccao.outro_produto, faccao.get_situacao_trabalho_display(), faccao.get_situacao_remuneracao_display(), faccao.get_voce_prefere_display(), faccao.qual_seu_sonho_no_setor])
        wb.save(response)
        return response
    else:
        return HttpResponse('Acesso negado')


@login_required
def cadastrar_faccao_legal(request):
    if request.method == 'POST':
        form = Faccao_Legal_Form(request.POST)
        if form.is_valid():
            faccao = form.save()
            faccao.user = request.user
            faccao.save()            
            try:
                if request.POST['cadastrar_empresa'] == 'on':
                    messages.success(request, 'Facção cadastrada com sucesso! Agora efetue o cadastro da empresa.')
                    return redirect('empreendedor:cadastrar_empresa')
            except:
                messages.success(request, 'Facção cadastrada com sucesso!')
            return redirect('empreendedor:faccao_legal')
    else:
        form = Faccao_Legal_Form(initial={'user': request.user.id})
    context = {
        'titulo': 'Sala do Empreendedor - Cadastrar Facção Legal',
        'titulo_pag':'Facção Legal',
        'form': form,
    }
    return render(request, 'sala_do_empreendedor/cadastro_faccao_legal.html', context)

import json
from django.http import JsonResponse
import re

def checkProfissao(request):    
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        id = data.get('id')
        try:
            profissao = Profissao.objects.get(id=id)
        except:
            profissao = False
        if profissao:
            if profissao.escolaridade.id == 1 or profissao.escolaridade.id == 2 or profissao.escolaridade.id == 3 or profissao.escolaridade.id == 4:
                diploma = True
            else: 
                diploma = False
            response_data = {'exists': True, 'licenca_sanitaria': profissao.licenca_sanitaria, 'diploma': diploma, 'licenca_ambiental': profissao.licenca_ambiental}
        else:
            response_data = {'exists': False}

        return JsonResponse(response_data)
    return JsonResponse({})

def checkCPF(request):    
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        cpf = data.get('cpf')
        try:
            pessoa = Pessoa.objects.get(cpf=re.sub(r'[^0-9]', '', cpf))
        except:
            pessoa = False
        print(cpf, pessoa)
        if pessoa:
            response_data = {'exists': True, 'nome': pessoa.nome}
        else:
            response_data = {'exists': False}

        return JsonResponse(response_data)
    return JsonResponse({})

def checkCNPJ(request):    
    if request.method == 'POST':
        data = json.loads(request.body.decode('utf-8'))
        cnpj = data.get('cnpj')
        try:
            empresa = Empresa.objects.get(cnpj=re.sub(r'[^0-9]', '', cnpj))
        except:
            empresa = False
        if empresa:
            response_data = {'exists': True}
        else:
            response_data = {'exists': False}

        return JsonResponse(response_data)
    return JsonResponse({})

def cadastro_fornecedores_e_compras_publicas(request):
    context = {
        'titulo': 'Cadastro de Fornecedores e Compras Públicas',
        'titulo_pag':'Cadastro de Fornecedores e Compras Públicas',
    }
    return render(request, 'sala_do_empreendedor/cadastro_fornecedores_e_compras.html', context)

def iss_autonomos(request):
    context = {
        'titulo': 'Sala do Empreendedor - ISS ou Autônomos',
        'titulo_pag':'ISS ou Autônomo',
    }
    return render(request, 'sala_do_empreendedor/iss_autonomos.html', context)

def legislacao(request):
    context = {
        'titulo': 'Sala do Empreendedor - Legislação',
        'titulo_pag':'Legislação',
    }
    return render(request, 'sala_do_empreendedor/legislacao.html', context)

def oportunidade_de_negocios(request):
    context = {
        'titulo': 'Sala do Empreendedor - Oportunidade de Negócios',
        'titulo_pag':'Oportunidade de Negócios',
    }
    return render(request, 'sala_do_empreendedor/oportunidade_de_negocios_index.html', context)

def oportunidade_de_negocios_dados(request):
    context = {
        'titulo': 'Sala do Empreendedor - Oportunidade de Negócios',
        'titulo_pag':'Oportunidade de Negócios',
    }
    return render(request, 'sala_do_empreendedor/oportunidade_de_negocios_graficos.html', context)


def novas_oportunidades(request):
    if request.method == 'POST':
        form = Form_Novas_Oportunidades(request.POST)
        if form.is_valid():            
            cadastro_artesao=form.save()
            if request.user.is_authenticated:
                cadastro_artesao.user_register = request.user
            cadastro_artesao.save()
            messages.success(request, 'Formulário enviado com sucesso!')
            return redirect('empreendedor:reuniao_sebrae')
        else:
            print(form.errors)
            messages.error(request, 'Erro ao enviar formulário. Verifique os campos e tente novamente.')
    else:
        form = Form_Novas_Oportunidades()
    context ={
        'titulo': 'Sala do Empreendedor - Oportunidade de Negócios',
        'form': form,        
    }
    return render(request, 'sala_do_empreendedor/form_novas_oportunidades.html', context)

@login_required
def export_novas_oportunidades(request):
    if request.user.is_superuser:
        response = HttpResponse(content_type='application/ms-excel')
        data = datetime.now().strftime('%d-%m-%Y')
        response['Content-Disposition'] = f'attachment; filename="novas_oportunidades-{data}.xls"'
        wb = Workbook()
        ws = wb.active
        ws.title = "Facções"
        ws.append(['CPF', 'Telefone', 'Email', 'Atividade Manual', 'Tipo de Costura', 'Tipo de Produção de Alimentos', 
                   'Tipo de Produção de Bebidas', 'Renda Representação', 'Motivo Não Comercialização', 'Renda Mensal', 
                   'Possui Empresa', 'Comercialização Produto', 'CEP Negócio', 'Alavancar Negócio', 'Usuário que Cadastrou', 
                   'Data de Cadastro'])
        
        novas_oportunidades = Novas_Oportunidades.objects.all()

        for oportunidade in novas_oportunidades:
            atividade_manual = ', '.join([str(item) for item in oportunidade.atividade_manual.all()])
            tipo_costura = ', '.join([str(item) for item in oportunidade.tipo_costura.all()])
            tipo_producao_alimentos = ', '.join([str(item) for item in oportunidade.tipo_producao_alimentos.all()])
            tipo_producao_bebidas = ', '.join([str(item) for item in oportunidade.tipo_producao_bebidas.all()])

            ws.append([
                oportunidade.cpf,
                oportunidade.telefone,
                oportunidade.email,
                atividade_manual,
                tipo_costura,
                tipo_producao_alimentos,
                tipo_producao_bebidas,
                oportunidade.get_renda_representacao_display(),
                oportunidade.get_motivo_nao_comercializacao_display(),
                oportunidade.get_renda_mensal_display(),
                oportunidade.get_possui_empresa_display(),
                oportunidade.get_comercializacao_produto_display(),
                oportunidade.cep_negocio,
                oportunidade.alavancar_negocio,
                oportunidade.user_register.username if oportunidade.user_register else '',
                oportunidade.dt_register.strftime('%d-%m-%Y'),
            ])
        wb.save(response)
        return response
    else:
        return HttpResponse('Acesso negado')



def redirecionamento_novas_oportunidades(request):
    context={
        'titulo': 'Sala do Empreendedor - Oportunidade de Negócios',
    }
    return render( request, 'novas_oportunidades_red.html', context)
def vitrine_virtual(request):
    registros=Registro_no_vitrine_virtual.objects.all().order_by('?')
    empresa_e_produtos=[]
    for registro in registros:
        produtos=Produto.objects.filter(rg_vitrine=registro)
        # print(produtos)
        if registro.logo:
            try:
                empresa_e_produtos.append({"empresa": registro.empresa, "logo": str(registro.logo.url), "produtos": produtos})
            except:
                pass
        else:
            try:
                empresa_e_produtos.append({"empresa": registro.empresa, "logo": None, "produtos": produtos})
            except:
                pass
            
    paginator = Paginator(empresa_e_produtos, 100) 
    page = request.GET.get('page')
    registros_paginated = paginator.get_page(page)

    context = {
        'titulo': 'Sala do Empreendedor - Vitrine Virtual',
        'titulo_pag': 'Vitrine Virtual',
        'registros': registros_paginated,
    }
    return render(request, 'sala_do_empreendedor/vitrine-virtual/vitrine.html', context)

def quero_ser_mei(request):
    context = {
         'titulo': 'Sala do Empreendedor - Quero ser MEI',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei.html', context)

#views do QUERO SER MEI
def por_que_ser_mei(request):
    context = {
         'titulo': 'Sala do Empreendedor',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/por-que-ser-mei.html', context)
def o_que_precisa_saber_para_ser_mei(request):
    context = {
         'titulo': 'Sala do Empreendedor - O que você precisa saber para ser MEI',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/o-que-voce-precisa-saber-antes-de-se-tornar-um-mei.html', context)
def jornada_empreendedora(request):
    context = {
         'titulo': 'Sala do Empreendedor - Jornada Empreendedora',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/jornada-empreendedora.html', context)
def documentosNecessarios(request):
    context = {
         'titulo': 'Sala do Empreendedor - Documentos Necessários',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/documentosNecessarios.html', context)

def quaisAsOcupacoesQuePodemSerMei(request):
    context = {
        'titulo':'Sala do Empreendedor - Quais as ocupações que podem ser MEI',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/quaisAsOcupacoesQuePodemSerMei.html', context)

def dicasDeSegurancaDaVigilanciaSanitaria(request):
    context = {
        'titulo':'Sala do Empreendedor - Dicas de Segurança da Vigilância Sanitária',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/dicasDeSegurancaDaVigilanciaSanitaria.html', context)

def dicasDeSegurançaDoCorpoDeBombeiros(request):
    context = {
        'titulo':'Sala do Empreendedor - Dicas de Segurança do Corpo de Bombeiros',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/dicasDeSegurançaDoCorpoDeBombeiros.html', context)

def dicasDeMeioAmbiente(request):
    context = {
        'titulo':'Sala do Empreendedor - Dicas de Meio Ambiente',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/dicasDeMeioAmbiente.html', context)

def prepareSe(request):
    context = {
        'titulo':'Sala do Empreendedor - Quero Ser MEI - Prepare-se',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/prepareSe.html', context)

def transportadorAutonomoDeCargas(request):
    context = {
        'titulo':'Sala do Empreendedor - Quero Ser MEI - Transportador Autônomo de Cargas',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/transportadorAutonomoDeCargas.html', context)

def direitosEObrigacoes(request):
    context = {
        'titulo':'Sala do Empreendedor - Quero Ser MEI - Direitos e Obrigações',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/direitosEObrigacoes.html', context)

def registrocadastur(request):
    context = {
         'titulo': 'Sala do Empreendedor - Quero Ser MEI - Registro Cadastur',
    }
    return render(request, 'sala_do_empreendedor/quero-ser-mei/registrocadastur.html', context)

def ja_sou_mei(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já sou MEI',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/ja-sou-mei.html', context)

def emissaoDeComprovante(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Emissão de Comprovante',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/emissaoDeComprovante.html', context)

def atualizacaoCadastral(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Atualização Cadastral',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/atualizacaoCadastral.html', context)

def capacita(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Capacita',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/capacita.html', context)

def notaFiscal(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Nota Fiscal',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/notaFiscal.html', context)

def relatorioMensal(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Relatório Mensal',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/relatorioMensal.html', context)

def pagamentoDeContribuicaoMensal(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Pagamento de Contribuição Mensal',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/pagamentoDeContribuicaoMensal.html', context)

def solucoesFinanceiras(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Soluções Financeiras',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/solucoesFinanceiras.html', context)

def certidoesEComprovantes(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Certidões e Comprovantes',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/certidoesEComprovantes.html', context)

def declaracaoAnualDeFaturamento(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Declaração Anual de Faturamento',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/declaracaoAnualDeFaturamento.html', context)

def dispensaDeAlvara(request):
    context = {
         'titulo': 'Sala do Empreendedor - Já  Sou Mei - Dispensa de Alvará',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/jaSouMei/dispensaDeAlvara.html', context)

def abertura_de_empresa(request):
    context = {
         'titulo': 'Sala do Empreendedor - Abertura de Empresa',
        # 'titulo': apps.get_app_config('financas').verbose_name,
    }
    return render(request, 'sala_do_empreendedor/abertura-de-empresa.html', context)

@login_required()
def requerimento_iss(request):
    # if True:
    #     return render(request, 'em-construcao.html', {})
    if request.method == 'POST':
        form = Criar_Processo_Form(request.POST, request.FILES)
        form_iss = Processo_ISS_Form(request.POST, request.FILES)
        if form.is_valid() and form_iss.is_valid():
            processo = form.save(commit=False)
            processo.tipo_processo = Tipo_Processos.objects.get(id=1)
            processo.solicitante = request.user
            processo.status = 'ae'
            processo.save()
            processo_iss = form_iss.save(commit=False)
            processo_iss.processo = processo
            processo_iss.save() 
            messages.success(request, 'Processo iniciado. Agora, envie os documentos necessários.')
            andamento = Andamento_Processo_Digital(
                processo=processo,              
                status='ae',
                observacao = 'Processo iniciado. Aguardando envio de documentos.',
                servidor = None
            )
            andamento.save()
            # status=Processo_Status_Documentos_Anexos(processo=processo)
            # status.save()
            send_email_for_create_process(processo, andamento)
            return redirect('empreendedor:requerimento_iss_doc', n_protocolo=processo.n_protocolo) 
        else:
            print(form.errors)
    else:
        form = Criar_Processo_Form(initial={'tipo_processo': 1, 'solicitante': request.user.id})
        form_iss = Processo_ISS_Form(initial={'solicitante': request.user.id})
    context = {
        'titulo': 'Sala do Empreendedor - Requerimento de ISS',
        'form': form,
        'form_iss': form_iss,
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/cadastro_processo.html', context)

def requerimento_documentos(request, n_protocolo):
    processo=Processo_Digital.objects.get(n_protocolo=n_protocolo)
    requerimento_iss = RequerimentoISS.objects.get(processo=processo)
    try:
        status = Processo_Status_Documentos_Anexos.objects.get(processo=processo)
    except:
        status = False
    if status:
        messages.warning(request, 'Processo já iniciado. Aguarde a avaliação dos documentos.')
        return redirect('empreendedor:andamento_processo', protocolo=n_protocolo)
    if request.method == 'POST':
        form = Criar_Processo_Docs_Form(request.POST, request.FILES)
        if form.is_valid():
            documentos = form.save(commit=False)
            documentos.processo = processo
            documentos.user_register = request.user
            try:                
                documentos.save()
                messages.success(request, 'Processo criado com sucesso!')
            except:
                messages.error(request, 'Error ao enviar os documentos. O nome dos arquivos anexados não devem conter acentos, cedilha ou caracteres especiais. Exemplo: ç, á, é, ã, õ, ô, ì, ò, ë, ù, ï, ü, etc.')
                return redirect('empreendedor:requerimento_iss_doc', protocolo=n_protocolo)
            andamento = Andamento_Processo_Digital(
                processo=processo,              
                status='nv',
                observacao = 'Processo criado. Aguardando avaliação do pedido.',
                servidor = None
            )
            andamento.save()
            processo.status = 'nv'
            processo.save()
            send_email_for_att_process(processo, andamento)
            if request.user.is_staff:
                return redirect('empreendedor:processos_digitais_admin')
            return redirect('empreendedor:listar_processos')
        else:
            print(form.errors)
    else:
        form = Criar_Processo_Docs_Form(initial={'processo': processo.id, 'user_register': request.user.id})
    # print(requerimento_iss.profissao.escolaridade.id)
    if requerimento_iss.profissao.escolaridade.id in [1,2,3,4]:
        # print('Tem diploma')
        not_diploma = False
    else:
        # print('Não tem diploma')
        not_diploma = True
    context = {
        'titulo': 'Sala do Empreendedor - Requerimento de ISS',
        'form': form, 
        'processo': processo,
        'requerimento': requerimento_iss,
        'not_diploma': not_diploma
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/cadastro_processo_doc.html', context)


def andamento_processo_iss(request, processo):
    requerimento_iss = RequerimentoISS.objects.get(processo=processo)
    andamentos = Andamento_Processo_Digital.objects.filter(processo=processo).order_by('-id')
    try:
        status_documentos = Processo_Status_Documentos_Anexos.objects.get(processo=processo)
    except:
        return redirect('empreendedor:requerimento_iss_doc', n_protocolo=processo.n_protocolo)
    context = {
        'titulo': 'Sala do Empreendedor - Andamento do Processo Online',
        'processo': processo,
        'andamentos': andamentos,
        'status_documentos': status_documentos,
        'requerimento': requerimento_iss,
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/andamento_processo_iss.html', context)

def andamento_processo_uniprofissional(request, processo):
    requerimento = RequerimentoISSQN.objects.get(processo=processo)
    andamentos = Andamento_Processo_Digital.objects.filter(processo=processo).order_by('-id')
    try:
        status_documentos = DocumentosPedido.objects.get(requerimento=requerimento)
    except:
        return redirect('empreendedor:requerimento_issqn_doc', n_protocolo=processo.n_protocolo)  
    context = {
        'titulo': 'Sala do Empreendedor - ISS Uniprofissional',
        'processo': processo,
        'andamentos': andamentos,
        'status_documentos': status_documentos,
        'requerimento': requerimento,
        
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/andamento_processo_uniprofissional.html', context)


@login_required()
def andamento_processo(request, protocolo):
    # messages.success(request, 'Processo encontrado.')
    processo = Processo_Digital.objects.get(n_protocolo=protocolo)
    print(processo.tipo_processo.id)
    if processo.tipo_processo.id == 1:
        return andamento_processo_iss(request, processo)
    elif processo.tipo_processo.id == 3:
        return andamento_processo_uniprofissional(request, processo)
    return redirect('empreendedor:listar_processos')

@login_required
def atualizar_documento_processo(request, protocolo, doc):
    try:
        processo = Processo_Digital.objects.get(n_protocolo = protocolo)
        if processo.tipo_processo.id == 1:
            status = Processo_Status_Documentos_Anexos.objects.get(processo = processo)
        elif processo.tipo_processo.id == 3:
            requerimento = RequerimentoISSQN.objects.get(processo = processo)
            status = DocumentosPedido.objects.get(requerimento = requerimento)
        try:
            status_doc = getattr(status, f"{doc}_status")
        except:
            messages.warning(request, 'Este documento não existe.')
            # TentativaBurla.objects.create(
            #     local_deteccao=f'empreendedor:atualizar_documento_processo -> /sala-do-empreendedor/processos-digitais/{protocolo}/att-doc/{doc}',
            #     user=request.user,
            #     ip_address=request.META.get('REMOTE_ADDR'),
            #     informacoes_adicionais=f'Possivel tentativa de acessar arquivo alterando url. Documento não encontrado: {doc}'
            # )
            return redirect('empreendedor:andamento_processo', protocolo=protocolo)
        if status_doc == '2':
            if request.method == 'POST':
                # Validar se o arquivo é válido antes de atualizar
                novo_documento = request.FILES[doc]
                # Faça a validação necessária, por exemplo, tipo de arquivo, tamanho, etc.
                        
                with transaction.atomic():
                    # Atualizar o documento
                    setattr(status, doc, novo_documento)
                    status.save()

                    # Mudar o status para 'Aguardando avaliação'
                    setattr(status, f"{doc}_status", '0')
                    status.save()

                # Muda o status para 'Aguardando avaliação'
                setattr(status, f"{doc}_status", '0')
                status.save()
                processo.save()
                # Verifica se todos os documentos estão aprovados para poder seguir com o processo
                if processo.tipo_processo.id == 1:
                    is_reprovado = (
                        status.rg_status == '2' or
                        status.comprovante_endereco_status == '2' or
                        status.diploma_ou_certificado_status == '2' or
                        status.licenca_sanitaria == '2' or
                        status.espelho_iptu_status == '2'
                    )
                elif processo.tipo_processo.id == 3:
                    is_reprovado = (
                        status.contrato_social_status == '2' or
                        status.carteira_orgao_classe_status == '2' or
                        status.alvara_localizacao_status == '2' or
                        status.informacoes_cadastrais_dos_empregados_status == '2' or
                        status.balanco_patrimonial_status == '2' or
                        status.dre_status == '2' or
                        status.balancete_analitico_status == '2' or
                        status.cnpj_copia_status == '2' or
                        status.profissionais_habilitados_status == '2' or
                        status.ir_empresa_status == '2' or
                        status.simples_nacional_status == '2'
                    )
                # Se não há documento reprovado, o processo pode seguir
                if not is_reprovado:
                    # Criasse então um novo andamento para o processo
                    Andamento_Processo_Digital.objects.create(
                        processo=processo,              
                        status='aa', 
                        observacao = 'Processo atualizado. Aguardando nova avaliação dos documentos.',
                        servidor = None
                    )            
                    processo.status = 'aa'
                    processo.save()
                    messages.success(request, 'Documento enviado com sucesso! Aguarde a nova avaliação dos documentos.')
                else:
                    messages.warning(request, 'Documento enviado com sucesso! Termine de atualizar os outros documentos.')
                return redirect('empreendedor:andamento_processo', protocolo=protocolo)
            context={
                'documento': doc
            }
            return render(request, 'sala_do_empreendedor/processos_digitais/att_documento.html', context)
        messages.warning(request, 'Este arquivo não necessita de alteração.')
        return redirect('empreendedor:andamento_processo', protocolo=protocolo)
    except Processo_Digital.DoesNotExist:
        messages.warning(request, 'Processo não encontrado.')
    except Processo_Status_Documentos_Anexos.DoesNotExist:
        messages.warning(request, 'Status de documentos não encontrado.')

    return redirect('empreendedor:andamento_processo', protocolo=protocolo)

def consultar_processos(request):
    context = {
        'titulo': 'Sala do Empreendedor - Consultar Processo Desenvolve NF',
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/consultar_processo.html', context)

@login_required()
def meus_processos(request):
    proecsesos = Processo_Digital.objects.filter(solicitante=request.user).order_by('-dt_solicitacao')    
    paginator = Paginator(proecsesos, 50)
    context = {
        'titulo': 'Sala do Empreendedor - Meus Processos Online',
        'processos': paginator.get_page(request.GET.get('page')),
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/listar_processos.html', context)

@login_required
def novo_processo(request):
    return render(request, 'sala_do_empreendedor/processos_digitais/novo_processo.html', {})

@login_required()
def pdde_index(request):
    cotext = {
        'titulo': 'Sala do Empreendedor - PDDE',
    }
    return render(request, 'sala_do_empreendedor/pdde/index.html', cotext)

@login_required()
def pdde_admin(request):
    cotext = {
        'titulo': 'Sala do Empreendedor - ADM - PDDE',
        'escolas': Escola.objects.all()
    }
    return render(request, 'sala_do_empreendedor/pdde/admin.html', cotext)

@login_required()
def pdde_index_escola(request):
    cotext = {
            'titulo': 'Sala do Empreendedor - PDDE Escola',
            'escolas': Escola.objects.filter(responsavel=request.user),
            # 'nome_da_escola': escola.nome,
        }
    return render(request, 'sala_do_empreendedor/pdde/index_escola.html', cotext)

@login_required()
def pdde_index_empresa(request):
    try:
        empresa = Empresa.objects.filter(user_register=request.user)
        if empresa.count() == 0:
            empresa=False
    except:
        empresa=False
    cotext = {
            'titulo': 'Sala do Empreendedor - PDDE Empresa',
            'solicitacoes': Solicitacao_de_Compras.objects.exclude(status='0'),
            'empresa': empresa
        }
    return render(request, 'sala_do_empreendedor/pdde/index_empresa.html', cotext)

@login_required()
def pdde_index_empresa_detalhe_solicitacao(request, id):
    solicitacao = Solicitacao_de_Compras.objects.get(id=id)
    itens = Item_Solicitacao.objects.filter(solicitacao_de_compra=solicitacao)
    if request.method == 'POST':
        if 'empresa' in request.POST:
            itens_valores=[]
            qnt=0
            for item in itens:
                try:
                    itens_valores.append([item.id, request.POST['proposta-'+str(item.id)]])
                    if request.POST['proposta-'+str(item.id)]!='0,00':
                        qnt+=1
                except:
                    pass
            if itens.count() == len(itens_valores):
                print(request.POST)
                empresa=Empresa.objects.get(id=int(request.POST['empresa']))
                if empresa.user_register == request.user:
                    proposta=Proposta.objects.create(
                        qnt_itens_proposta = qnt,
                        solicitacao_de_compra=solicitacao,
                        empresa=empresa,
                        previsao_entrega = request.POST['dt_previsao']
                    )
                    
                    for item_valor in itens_valores:
                        print(item_valor)
                        preco_=item_valor[1].replace(',', '')
                        preco=preco_.replace('.', '')
                        Proposta_Item.objects.create(
                            proposta=proposta,
                            item_solicitacao=Item_Solicitacao.objects.get(id=int(item_valor[0])),
                            preco = preco
                        )
                    messages.success(request, 'Proposta enviada com sucesso!')
                    return redirect('empreendedor:pdde_empresa')
            else:
                messages.warning(request, 'Preencha todos os campos de proposta corretamente.')
        else:
            messages.warning(request, 'Você precisia de uma empresa para enviar a proposta.')
    if solicitacao.status == '3' or solicitacao.status == '4' or solicitacao.status == '5':
        total_propostas = Proposta.objects.filter(solicitacao_de_compra=solicitacao).count()
    else:
        total_propostas = 0
    cotext = {
            'titulo': 'Sala do Empreendedor - PDDE Proposta',
            'solicitacao': solicitacao,
            'itens': itens,
            'total_propostas': total_propostas,
            'tipo': str(solicitacao.get_tipo_display()).lower(),
            'empresas': Empresa.objects.filter(user_register=request.user)
        }
    return render(request, 'sala_do_empreendedor/pdde/listar_itens_para_empresa.html', cotext)
   
@login_required()
def pdde_criar_escola(request):
    if request.method == 'POST':
        form = Escola_Form(request.POST)
        if form.is_valid():
            escola = form.save(commit=False)
            escola.responsavel = request.user
            escola.user_register = request.user
            escola.save()
            messages.success(request, 'Escola cadastrada com sucesso! Aguarde a validação de nossa equipe.')
            return redirect('empreendedor:pdde_escola')
        else:
            print('Erro ao cadastrar escola:', form.errors)            
    else:
        form = Escola_Form(initial={'responsavel': request.user.id})
    context = {
        'titulo': 'Sala do Empreendedor - PDDE Escola - Nova escola',
        'form': form
    }
    return render(request, 'sala_do_empreendedor/pdde/criar_escola.html', context)

@login_required()
def pdde_editar_escola(request, id):
    escola=Escola.objects.get(id=id)
    if request.method == 'POST':
        form = Escola_Form(request.POST, instance=escola)
        if form.is_valid():
            escola = form.save(commit=False)
            escola.user_register = request.user
            escola.save()
            messages.success(request, 'Escola cadastrada com sucesso!')
            return redirect('empreendedor:pdde_admin')
    else:
        form = Escola_Form(instance=escola)
    context = {
        'titulo': 'Sala do Empreendedor - PDDE Escola - Editar escola',
        'form': form
    }
    return render(request, 'sala_do_empreendedor/pdde/criar_escola.html', context)

@login_required()
def pdde_criar_solicitacao_de_compra(request, id):
    escola=Escola.objects.get(id=id)
    if request.method == 'POST':
        form = Solicitacao_de_Compras_Form(request.POST)
        if form.is_valid():
            solicitacao=form.save()
            solicitacao.escola=escola
            solicitacao.save()
            messages.success(request, 'Solicitação cadastrada com sucesso!')
            return redirect('empreendedor:pdde_criar_item_solicitacao', id=solicitacao.id)
    else:
        form = Solicitacao_de_Compras_Form(initial={'escola': escola.id})
    context = {
        'titulo': 'Sala do Empreendedor - PDDE Escola - Nova solicitação',
        'escola': escola,
        'form': form
    }
    return render(request, 'sala_do_empreendedor/pdde/criar_solitacao_de_compra.html', context)

@login_required()
def pdde_confirmar_pagamento(request, hash):
    contrato = Contrato_de_Servico.objects.get(hash = hash)
    if request.method == 'POST':
        if contrato.solicitacao_referente.escola.responsavel == request.user:
            contrato.solicitacao_referente.status = '6'
            contrato.solicitacao_referente.save()
            contrato.save()
            messages.success(request, 'Pagamento confirmado! Aguardando envio de nota fiscal.')
            return redirect('empreendedor:pdde_listar_solicitacoes', contrato.solicitacao_referente.id)
        else:
            messages.warning(request, 'Você não tem autorização para isso.')
            # TentativaBurla.objects.create(
            #     local_deteccao=f'empreendedor:pdde_confirmar_pagamento -> solicitacao -> {contrato.solicitacao_referente}',
            #     user=request.user,
            #     ip_address=request.META.get('REMOTE_ADDR'),
            #     informacoes_adicionais=f'Possivel tentativa de confirmar pagamento por url.'
            # )
    return render(request, 'sala_do_empreendedor/pdde/confirmar-pagamento.html')

@login_required()
def pdde_criar_itens_solicitacao(request, id):
    
    solicitacao=Solicitacao_de_Compras.objects.get(id=id)
    if solicitacao.status == '3':
        contrato = Contrato_de_Servico.objects.get(solicitacao_referente=solicitacao)
        return redirect('empreendedor:pdde_contratacao', hash=contrato.hash)
    elif solicitacao.status == '4':
        # print(solicitacao.get_status_display(), solicitacao.status)
        contrato = Contrato_de_Servico.objects.get(solicitacao_referente=solicitacao)
        return redirect('empreendedor:pdde_aguardando_execucao', hash=contrato.hash)
    elif solicitacao.status=='5':
        contrato = Contrato_de_Servico.objects.get(solicitacao_referente=solicitacao)
        return redirect('empreendedor:pdde_confirmar_pagamento', hash=contrato.hash)
    elif int(solicitacao.status)>=6:
        contrato = Contrato_de_Servico.objects.get(solicitacao_referente=solicitacao)
        return redirect('empreendedor:pdde_menu_escola', hash=contrato.hash)
    # soma={'menor_valor': 0, 'maior_valor': 0}
    if request.method == 'POST':
        response = PDDE_POST(request, solicitacao)
        if response == 'salvo':
            messages.success(request, 'Solicitação criada/iniciada com sucesso! Aguardando propostas.')
            return redirect('empreendedor:pdde_listar_solicitacoes', id=solicitacao.escola.id)
        elif response == 'escola-inativa':
            messages.warning(request, 'Sua escola ainda não foi aprovada pela equipe da Sala do Empreendedor. Aguarde a aprovação para poder criar solicitações.')
            return redirect('empreendedor:pdde_escola')
        elif response[0] == 'proposta-aceita':
            messages.success(request, 'Proposta aceita com sucesso! Aguardando a contratação da empresa.')            
            return redirect('empreendedor:pdde_contratacao', id=response[1].hash)
        itens=Item_Solicitacao.objects.filter(solicitacao_de_compra=solicitacao)
    # elif solicitacao.status != '0':
        # itens, soma = Item_E_Melhor_Proposta(solicitacao)
        # form = Criar_Item_Solicitacao(initial={'solicitacao_de_compra': solicitacao.id})
    else:
        form = Criar_Item_Solicitacao(initial={'solicitacao_de_compra': solicitacao.id})
        itens=Item_Solicitacao.objects.filter(solicitacao_de_compra=solicitacao)
    
    # soma={
    #         'menor_valor': '{:,.2f}'.format(soma['menor_valor']/100).replace('.', '##').replace(',', '.').replace('##', ','), 
    #         'maior_valor': '{:,.2f}'.format(soma['maior_valor']/100).replace('.', '##').replace(',', '.').replace('##', ',')
    #         }
    context = {
        'titulo': 'Sala do Empreendedor - PDDE Escola - Criar itens',
        'solicitacao': solicitacao,
        'itens': itens,
        'form': form,
        'lista_propostas': Listar_Proposta(id),
        # 'soma': soma
    }
    return render(request, 'sala_do_empreendedor/pdde/criar_itens_solicitacao.html', context)

def pdde_contratacao(request, hash):
    contrato = get_object_or_404(Contrato_de_Servico, hash=hash)
    if contrato.proposta_vencedora.empresa.user_register == request.user:
        contratado = True
    else:
        contratado = False
    if request.method == 'POST':
        if contratado == True:
            contrato.solicitacao_referente.status = '4'
            contrato.solicitacao_referente.save()
            messages.success(request, 'Contrato assinado com sucesso! Aguardando execução do serviço.')
            return redirect('empreendedor:pdde_empresa_menu', hash=contrato.hash)
    context = {
        'contrato': contrato,
        'contratado': contratado,
               }
    return render(request, 'sala_do_empreendedor/pdde/contratacao.html', context)

def pdde_nota_fiscal(request, hash):
    contrato = Contrato_de_Servico.objects.get(hash=hash)
    if request.method == 'POST':
        if contrato.solicitacao_referente.escola.responsavel == request.user:
            form = Contrato_NotaFiscal(request.POST, request.FILES, instance=contrato)
            if form.is_valid():
                contrato=form.save()
                contrato.solicitacao_referente.status = '7'
                contrato.solicitacao_referente.save()
                messages.success(request, 'Nota fiscal enviada com sucesso! Aguarde a avaliação da equipe da Sala do Empreendedor.')
                return redirect('empreendedor:pdde_empresa_menu', hash=contrato.hash )
    else:
        form = Contrato_NotaFiscal(instance=contrato)
    context={
        'contrato': contrato,
        'form': form
    }
    return render(request, 'sala_do_empreendedor/pdde/envio_de_nota_fiscal.html', context)

@login_required
def pdde_menu_escola(request, hash):
    contrato = Contrato_de_Servico.objects.get(hash=hash)
    if not contrato.solicitacao_referente.escola.responsavel == request.user:
        messages.warning(request, 'Você não tem autorização para acessar essa área.')
        return redirect('empreendedor:pdde_escola')
    context = {
        'contrato': contrato
    }
    return render(request, 'sala_do_empreendedor/pdde/menu-escola.html', context)

@login_required
def pdde_avaliar_servico(request, hash):
    contrato = Contrato_de_Servico.objects.get(hash=hash)
    if request.method == 'POST':
        if contrato.solicitacao_referente.escola.responsavel == request.user:
            form = Contrato_Avaliacao(request.POST, instance=contrato)
            if form.is_valid():
                contrato.solicitacao_referente.status = '8'
                contrato.solicitacao_referente.save()
                messages.success(request, 'Serviço avaliado com sucesso! Solicitação concluída!')
                return redirect('empreendedor:pdde_empresa_menu', hash=contrato.hash )
    else:
        form = Contrato_Avaliacao(instance=contrato)
    context={
        'form': form,
        'contrato': contrato
    }
    return render(request, 'sala_do_empreendedor/pdde/avaliar-servico.html', context)

def pdde_menu_opcoes_empresa(request, hash):
    contrato = Contrato_de_Servico.objects.get(hash=hash)
    context={
        'contrato': contrato
    }
    return render(request, 'sala_do_empreendedor/pdde/menu-empresa.html', context)

def pdde_aguardando_execucao(request, hash):
    contrato = Contrato_de_Servico.objects.get(hash=hash)
    if request.method == 'POST':
        if contrato.solicitacao_referente.escola.responsavel == request.user:
            contrato.solicitacao_referente.status = '5'
            contrato.solicitacao_referente.save()
            messages.success(request, 'Execução do serviço confirmada! Aguardando confirmação do pagamento.')
            return redirect('empreendedor:pdde_confirmar_pagamento', hash=contrato.hash)
    context={
        'contrato': contrato
    }
    return render(request, 'sala_do_empreendedor/pdde/aguardando_execucao.html', context)


def listar_proposta_para_o_item(request, id, id_item):
    item = Item_Solicitacao.objects.get(id=id_item)
    propostas = Proposta_Item.objects.filter(item_solicitacao=item).order_by('preco')
    context = {
        'titulo': 'Sala do Empreendedor - PDDE Escola - Propostas',
        'item': item,
        'propostas': propostas,
    }
    return render(request, 'sala_do_empreendedor/pdde/listar_propostas.html', context)

def pdde_criar_itens_solicitacao_fetch(request):
    try:
        if request.method == 'POST':
            print(request.POST)
            solicitacao = Solicitacao_de_Compras.objects.get(id=request.POST.get('solicitacao_de_compra'))
            if request.user == solicitacao.escola.responsavel or request.user.is_staff:
                item = Item_Solicitacao(
                    solicitacao_de_compra=solicitacao,
                    nome=request.POST.get('nome'),  
                    quantidade=request.POST.get('quantidade'),
                    unidade=request.POST.get('unidade'),
                    descricao=request.POST.get('descricao'),
                )
                item.save()
                itens = Item_Solicitacao.objects.filter(solicitacao_de_compra=solicitacao)
                itens_json = []
                for item in itens:
                    itens_json.append({
                        'id': item.id,
                        'nome': item.nome,
                        'quantidade': item.quantidade,
                        'unidade': item.unidade,
                        'descricao': item.descricao,

                    })
                return JsonResponse(itens_json, safe=False)
        return JsonResponse({})
    except Exception as E:
        print(E)
        return JsonResponse({'error': 'Ocorreu um erro no servidor'}, status=500)

def pdde_remover_item_solicitacao_featch(request):
    try:
        if request.method == 'POST':
            print(request.POST)
            item = Item_Solicitacao.objects.get(id=request.POST.get('id'))
            solicitacao = item.solicitacao_de_compra
            if request.user == solicitacao.escola.responsavel or request.user.is_staff:
                item.delete()
                itens = Item_Solicitacao.objects.filter(solicitacao_de_compra=solicitacao)
                itens_json = []
                for item in itens:
                    itens_json.append({
                        'id': item.id,
                        'nome': item.nome,
                        'quantidade': item.quantidade,
                        'unidade': item.unidade,
                        'descricao': item.descricao,

                    })
                return JsonResponse(itens_json, safe=False)
        return JsonResponse({})
    except Exception as E:
        print(E)
        return JsonResponse({'error': 'Ocorreu um erro no servidor'}, status=500)   

@login_required
def pdde_listar_solicitacoes(request, id):
    try:
        escola=Escola.objects.get(id=id)
        if escola.responsavel != request.user and request.user.is_staff == False:
            messages.warning(request, 'Você não possui autorização para acessar essa página!')
            return redirect('empreendedor:pdde_index')
    except:
        messages.warning(request, 'Você não possui autorização para acessar essa página!')
        return redirect('empreendedor:pdde_index')
    solicitacoes=Solicitacao_de_Compras.objects.filter(escola=escola)
    context = {
        'titulo': 'Sala do Empreendedor - PDDE Escola - Listar solicitações',
        'escola': escola,
        'solicitacoes': solicitacoes
    }
    return render(request, 'sala_do_empreendedor/pdde/listar_solicitacoes.html', context)
 
@login_required
def requerimento_ISSQN(request):
    context = {
        'titulo': 'Sala do Empreendedor - Requerimento de ISSQN',
    }
    return render(request, 'sala_do_empreendedor/em-construcao.html', context)    
    # if request.method == 'POST':        
    #     form = Criar_Processo_Form(request.POST, request.FILES)
    #     form_req = RequerimentoISSQNForm(request.POST)
    #     documentos_pedido_form = DocumentosPedidoForm(request.POST, request.FILES)
    #     if form.is_valid() and form_req.is_valid() and documentos_pedido_form.is_valid():
    #         processo = form.save()
    #         requerimento = form_req.save()
    #         requerimento.processo = processo
    #         requerimento.save()
    #         documentos_pedido = documentos_pedido_form.save(commit=False)
    #         documentos_pedido.requerimento = requerimento
    #         try:
    #             documentos_pedido.save()                
    #         except:
    #             messages.error(request, 'Error ao enviar os documentos. O nome dos arquivos anexados não devem conter acentos, cedilha ou caracteres especiais. Exemplo: ç, á, é, ã, õ, ô, ì, ò, ë, ù, ï, ü, etc.')
    #             andamento = Andamento_Processo_Digital(
    #                 processo=processo,              
    #                 status='ae',
    #                 observacao = 'Processo iniciado. Aguardando envio de documentos.',
    #                 servidor = None
    #             )
    #             return redirect('empreendedor:requerimento_issqn_doc', n_protocolo=processo.n_protocolo)    
    #         andamento = Andamento_Processo_Digital(
    #             processo=processo,              
    #             status='aa',
    #             observacao = 'Processo iniciado. Aguardando avaliação de documentos.',
    #             servidor = None
    #         )
    #         andamento.save()
            
    #         #EL AQUI
    #         api = ApiProtocolo()
    #         data_processo = datetime.now()
    #         pessoa = Pessoa.objects.get(user=processo.solicitante)
    #         parametros = {
    #             'numeroDocumentoJuridico': str(pessoa.cpf),
    #             'nomePessoa': pessoa.nome,
    #             'numeroProcesso': str(data_processo.year),
    #             'anoProcesso': str(data_processo.year),
    #             'dataProcesso': data_processo.strftime('%Y-%m-%d %H:%M:%S'),
    #             'resumoEcm': f'Processo de Requerimento de ISS iniciado Desenvolve NF. Link para acompanhamento: https://desenvolve.novafriburgo.rj.gov.br/sala-do-empreendedor/processos-digitais/{processo.n_protocolo}/'
    #         }
    #         print(parametros)            
    #         print('Resposta', api.cadastrarProcesso(parametros))
    #         # api.cadastrarPessoa(parametros)
    #         messages.success(request, 'Processo criado. Aguardando avaliação de documentos.')
    #         send_email_for_create_process(processo, andamento)
    #         return redirect('empreendedor:andamento_processo', protocolo=processo.n_protocolo)
    #     else:
    #         print(form.errors)
    #         print(form_req.errors)
    #         print(documentos_pedido_form.errors)
    # else:
    #     form = Criar_Processo_Form(initial={'tipo_processo': 3, 'solicitante': request.user.id})
    #     form_req = RequerimentoISSQNForm()
    #     documentos_pedido_form = DocumentosPedidoForm()

    # context={
    #     'form': form,
    #     'form_req': form_req,
    #     'documentos_pedido_form': documentos_pedido_form,
    #     'titulo': 'Requerimento para ISSQN como Sociedade Uniprofissional'
    # }
    # return render(request, 'sala_do_empreendedor/processos_digitais/uniprofissional/index.html', context)

@login_required
def requerimento_ISSQN_doc(request, n_protocolo):
    processo = Processo_Digital.objects.get(n_protocolo=n_protocolo)
    requerimento = RequerimentoISSQN.objects.get(processo=processo)
    if request.method == 'POST':        
        documentos_pedido_form = DocumentosPedidoForm(request.POST, request.FILES)
        if documentos_pedido_form.is_valid():
            documentos_pedido = documentos_pedido_form.save(commit=False)
            documentos_pedido.requerimento = requerimento
            try:
                documentos_pedido.save()
            except:
                messages.error(request, 'Error ao enviar os documentos. O nome dos arquivos anexados não devem conter acentos, cedilha ou caracteres especiais. Exemplo: ç, á, é, ã, õ, ô, ì, ò, ë, ù, ï, ü, etc.')
                andamento = Andamento_Processo_Digital(
                    processo=processo,              
                    status='ae',
                    observacao = 'Processo iniciado. Aguardando envio de documentos.',
                    servidor = None
                )
                return redirect('empreendedor:requerimento_issqn_doc', n_protocolo=processo.n_protocolo)    
            andamento = Andamento_Processo_Digital(
                processo=processo,              
                status='aa',
                observacao = 'Processo iniciado. Aguardando avaliação de documentos.',
                servidor = None
            )
            andamento.save()
            messages.success(request, 'Documentos enviados com sucesso! Processo criado.')
            return redirect('empreendedor:andamento_processo', protocolo=processo.n_protocolo)
        else:
            print(documentos_pedido_form.errors)
    else:
        documentos_pedido_form = DocumentosPedidoForm()

    context={
        'documentos_pedido_form': documentos_pedido_form,
        'titulo': 'Requerimento para ISSQN como Sociedade Uniprofissional'
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/uniprofissional/documentos.html', context)


def atualizar_todo_dia(request):
    hoje = timezone.now().date()
    solicitacao = Solicitacao_de_Compras.objects.filter(status='1')
    # print(solicitacao.count())
    for s in solicitacao:
        # print('---------------------------')
        # print('model: ', s.dt_envio_propostas, ' today: ', hoje)
        if s.dt_envio_propostas == hoje:
            s.status = '2'
            s.save()
    return HttpResponse('ok')


def alimentar_oportunidades(request):
    #atividade manual
    atividades_cadastrar = [
        'Bordado', 'Costura', 'Pintura', 'Tapeçaria', 'Esculturas', 
        'Marcenaria', 'Produção de Alimentos', 'Produção de Bebidas', 
        'Decoração de festas/eventos', 'Fotografia/Filmagens/Edição', 
        'Produção Cultural/Musical', 'Outra', 'Atualmente não tenho nenhuma atividade manual'
    ]
    #tipo costura;
    tipo_costura_cadastrar = ['Sob medida', 'Facção', 'Criativa', 
                    'Roupas de Boneca', 'Infantil', 'Outra']
    
    #tipo producao alimentos
    tipo_alimentos_cadsatrar = ['Doce', 'Pães/bolos', 'Temperos', 
                                'Conservas', 'Queijos', 'Geleias',
                                'Salgados', 'Produção de mel', 'Refeições',
                                'Chocolates', 'Produtos Dietéticos/orgânicos',
                                'Outro']
    # tipo producao bebidas
    tipo_bebidas_cadastrar = ['Cerveja','Sucos','Chás',
                              'Refrigerantes','Licos, cachaça','Vinhos',
                              'Café','Outro'                         
    ]
    
    for atividade in atividades_cadastrar:
        AtividadeManual.objects.create(descricao=atividade)
    
    for costura in tipo_costura_cadastrar:
        Tipo_Costura.objects.create(descricao=costura)
        
    for alimento in tipo_alimentos_cadsatrar:
        Tipo_Producao_Alimentos.objects.create(descricao=alimento)
    
    for bebida in tipo_bebidas_cadastrar:
        Tipo_Producao_Bebidas.objects.create(descricao=bebida)
        
    return HttpResponse('processo concluido')

def credito_facil(request):
    if request.method == 'POST':
        form = Form_Credito_Facil(request.POST)
        if form.is_valid():
            credito = form.save()
            if request.user.is_authenticated:
                credito.user_register = request.user
            credito.save()
            messages.success(request, 'Solicitação enviada com sucesso!')
            return redirect('empreendedor:index')
        else:
            print(form.errors)
            print('-------------------------------------------------------------')
    else:
        print('GET')
        if request.user.is_authenticated:
            form = Form_Credito_Facil(initial={'user_register': request.user.id})
        else:
            form = Form_Credito_Facil()
    context = {
        'form': form,
        'titulo': 'Sala do Empreendedor - Solicitação de Crédito Fácil'
    }
    return render(request, 'sala_do_empreendedor/credito_facil/emprestimo.html', context)


@login_required
def export_empresas_to_excel(request):
    if request.user.is_superuser:
        pass
    else:
        HttpResponseForbidden('Você não tem acesso a essa rota.')
    # Cria um Workbook (arquivo Excel)
    if request.method == 'POST':
        wb = Workbook()
        ws = wb.active
        dt_now = dt_now = timezone.now().strftime('%Y-%m-%d_%H-%M-%S')
        ws.title = f"empresas_{dt_now}"  # Define o título da planilha

        # Escreve os cabeçalhos das colunas
        ws.append(['CNPJ', 'Nome da empresa', 'Porte da empresa', 'Atividade', 'Outra atividade',
                'Ramo de atuação', 'Outro ramo', 'Telefone de contato', 'Whatsapp da empresa',
                'E-mail da empresa', 'Site da empresa', 'Descrição da empresa', 'Data de cadastro', 'Deseja receber noticias?'])

        # Obtém todas as empresas do banco de dados
        ramos_selecionados = request.POST.getlist('ramo')        
        empresas = Empresa.objects.filter(ramo__in=ramos_selecionados, receber_noticias=True)

        # Adiciona os dados das empresas ao arquivo Excel
        for empresa in empresas:
            ws.append([empresa.cnpj, empresa.nome, empresa.porte.porte  , ', '.join(str(atividade) for atividade in empresa.atividade.all()),
                    empresa.outra_atividade, ', '.join(str(ramo) for ramo in empresa.ramo.all()), empresa.outro_ramo,
                    empresa.telefone, empresa.whatsapp, empresa.email, empresa.site, empresa.descricao,
                    empresa.dt_register, empresa.receber_noticias])

        # Define o nome do arquivo e o tipo de conteúdo da resposta HTTP
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="empresas_{dt_now}.xlsx"'

        # Salva o Workbook na resposta HTTP
        wb.save(response)

        return response
    context = {
        'form' : FormRamos()
    }
    return render(request, 'sala_do_empreendedor/admin/mapeamento_filtrar.html', context)


def importar_empresas(request):
    # if request.method == 'POST' and request.FILES['excel_file']:
    #     excel_file = request.FILES['excel_file']
    #     # Carregar o arquivo Excel em um DataFrame do pandas
    #     df = pd.read_excel(excel_file)

    #     # Iterar sobre as linhas do DataFrame e criar ou atualizar objetos Empresa
    #     for index, row in df.iterrows():
    #         cnpj = validate_CNPJ(str(row['Número do CNPJ']))
    #         nome =  row['Nome Fantasia']
    #         porte = 
    #         atividade =
    #         outra_atividade =
    #         ramo = 
    #         outro_ramo = 
    #         receber_noticias = 
    #         telefone = row['Telefone - Exemplo: (99) 9999-9999']
    #         whatsapp = row['Celular - Exemplo (99) 99999-9999']
    #         email = ''
    #         site
    #         descricao = f'Representante: {row['Nome do Representante Legal']}'
    #         user_register = User.objects.get(username='sistema')
    #         validacao = True
    #         cadastrada_como_fornecedo = True

    #         empresa, created = Empresa.objects.get_or_create(cnpj=cnpj)
    #         empresa.nome =
    #         empresa.telefone = row['Telefone - Exemplo: (99) 9999-9999']
    #         empresa.whatsapp = row['Celular - Exemplo (99) 99999-9999']
    #         empresa.porte = row['Porte da Empresa']
    #         empresa.save()

    #     return render(request, 'importar_sucesso.html')
    # else:
        return render(request, 'importar_form.html')

def natal_artesao(request):
    if request.method == 'POST':
        form = Form_Natal_Artesao(request.POST)
        form_2 = Form_Novas_Oportunidades(request.POST)
        if form.is_valid() and form_2.is_valid():            
            
            cadastro_artesao=form_2.save()
            cadastro_natal=form.save()
            if request.user.is_authenticated:
                cadastro_artesao.user_register = request.user
                cadastro_natal.user_register = request.user
            cadastro_artesao.save()
            cadastro_natal.save()
            messages.success(request, 'Formulário enviado com sucesso!')
            return render(request, 'sala_do_empreendedor/form_novas_oportunidades_natal _sucesso.html')
        else:
            print(form.errors)
            messages.error(request, 'Erro ao enviar formulário. Verifique os campos e tente novamente.')
    else:
       form = Form_Natal_Artesao()
       form_2 = Form_Novas_Oportunidades
    context ={
        'titulo': 'Sala do Empreendedor - Oportunidade de Negócios',
        'form': form,        
        'form_2': form_2
    }
    return render(request, 'sala_do_empreendedor/form_novas_oportunidades_natal.html', context)

def checkCPFArtesao(request):
    if request.method == 'POST':
        print('POST', print(request.POST))
        if Novas_Oportunidades.objects.filter(cpf=request.POST.get('cpf')).exists():
            return JsonResponse({'valid': False})
    return JsonResponse({'valid': False})

def imprimir_documentos(request, protocolo):
    processo = get_object_or_404(Processo_Digital, n_protocolo=protocolo)
    requerimento = get_object_or_404(RequerimentoISSQN, processo=processo)
    documentos = DocumentosPedido.objects.filter(requerimento=requerimento)
    
    context = {
        'titulo': 'Imprimir Documentos - Sala do Empreendedor',
        'processo': processo,
        'requerimento': requerimento,
        'documentos': documentos,
        'now': timezone.now(),
    }
    return render(request, 'sala_do_empreendedor/processos_digitais/uniprofissional/imprimir_documentos.html', context)