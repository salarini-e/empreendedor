from django.shortcuts import render, redirect
from .forms import *
from autenticacao.models import Pessoa
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.apps import apps
from .models import *
from settings.decorators import group_required
from django.contrib.auth.models import Group
from datetime import datetime
from django.db.models import Case, When, Count, Sum, IntegerField

from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime, timedelta
from django.utils import timezone
from .custom import month_translate
from django.db import connection, models

STATUS_CHOICES=(
        ('0','Novo'),
        ('1','Aguardando'),
        ('2','Em execução'),
        ('f','Finalizado')
    )
PRIORIDADE_CHOICES=(
        ('0','Normal'),
        ('1','Moderada'),
        ('2','Urgente'),
    )

@group_required('os_acesso')
def index(request):
    return render(request, 'os_index.html')
# from .functions import get_os_stats

# def relatorio_mensal(request, month, year):
#     stats = get_os_stats(month, year)
    
#     context = {
#         'os_by_bairro': stats['os_by_bairro'],
#         'total_os': stats['total_os'],
#         'avg_os': stats['avg_os'],
#         'median_os': stats['median_os'],
#         'points_by_bairro': stats['points_by_bairro'],
#         'difference': stats['difference']
#     }
    
#     return render(request, 'relatorio_mensal.html', context)

@login_required
@group_required('os_acesso')
def os_painel(request):    
    # meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    # bairros = OrdemDeServico.objects.values_list('bairro', flat=True).distinct()
    # data = []  

    # for bairro in bairros:
    #     total=OrdemDeServico.objects.filter(bairro=bairro, status__in=['0', '1', '2'], dt_solicitacao__year='2023').count()
    #     os_por_mes = OrdemDeServico.objects.filter(bairro=bairro, status__in=['0', '1', '2'], dt_solicitacao__year='2023').annotate(
    #         jan=Count('id', filter=models.Q(dt_solicitacao__month=1)),
    #         fev=Count('id', filter=models.Q(dt_solicitacao__month=2)),
    #         mar=Count('id', filter=models.Q(dt_solicitacao__month=3)),
    #         abr=Count('id', filter=models.Q(dt_solicitacao__month=4)),
    #         mai=Count('id', filter=models.Q(dt_solicitacao__month=5)),
    #         jun=Count('id', filter=models.Q(dt_solicitacao__month=6)),
    #         jul=Count('id', filter=models.Q(dt_solicitacao__month=7)),
    #         ago=Count('id', filter=models.Q(dt_solicitacao__month=8)),
    #         set=Count('id', filter=models.Q(dt_solicitacao__month=9)),
    #         out=Count('id', filter=models.Q(dt_solicitacao__month=10)),
    #         nov=Count('id', filter=models.Q(dt_solicitacao__month=11)),
    #         dez=Count('id', filter=models.Q(dt_solicitacao__month=12)),
    #     ).values('jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez')

    #     dt={'bairro': bairro, 'mes': os_por_mes, 'total': total}
    #     if not any(item['bairro'] == bairro for item in data):
    #         data.append(dt)
    # bairros = (
    #     OrdemDeServico.objects
    #     .filter(status__in=['0', '1', '2'], dt_solicitacao__year='2023')
    #     .values('bairro')
    #     .annotate(
    #         total=Count('id'),
    #         jan=Sum(models.Case(models.When(dt_solicitacao__month=1, then=1), default=0, output_field=models.IntegerField())),
    #         fev=Sum(models.Case(models.When(dt_solicitacao__month=2, then=1), default=0, output_field=models.IntegerField())),
    #         mar=Sum(models.Case(models.When(dt_solicitacao__month=3, then=1), default=0, output_field=models.IntegerField())),
    #         abr=Sum(models.Case(models.When(dt_solicitacao__month=4, then=1), default=0, output_field=models.IntegerField())),
    #         mai=Sum(models.Case(models.When(dt_solicitacao__month=5, then=1), default=0, output_field=models.IntegerField())),
    #         jun=Sum(models.Case(models.When(dt_solicitacao__month=6, then=1), default=0, output_field=models.IntegerField())),
    #         jul=Sum(models.Case(models.When(dt_solicitacao__month=7, then=1), default=0, output_field=models.IntegerField())),
    #         ago=Sum(models.Case(models.When(dt_solicitacao__month=8, then=1), default=0, output_field=models.IntegerField())),
    #         set=Sum(models.Case(models.When(dt_solicitacao__month=9, then=1), default=0, output_field=models.IntegerField())),
    #         out=Sum(models.Case(models.When(dt_solicitacao__month=10, then=1), default=0, output_field=models.IntegerField())),
    #         nov=Sum(models.Case(models.When(dt_solicitacao__month=11, then=1), default=0, output_field=models.IntegerField())),
    #         dez=Sum(models.Case(models.When(dt_solicitacao__month=12, then=1), default=0, output_field=models.IntegerField()))
    #     )
    # )

    # data = []
    # for bairro in bairros:
    #     bairro_data = {
    #         'bairro': bairro['bairro'],
    #         'mes': {
    #             'jan': bairro['jan'],
    #             'fev': bairro['fev'],
    #             'mar': bairro['mar'],
    #             'abr': bairro['abr'],
    #             'mai': bairro['mai'],
    #             'jun': bairro['jun'],
    #             'jul': bairro['jul'],
    #             'ago': bairro['ago'],
    #             'set': bairro['set'],
    #             'out': bairro['out'],
    #             'nov': bairro['nov'],
    #             'dez': bairro['dez'],
    #         },
    #         'total': bairro['total']
    #     }
    #     data.append(bairro_data)

    # query = '''
    # SELECT
    #     bairro,
    #     COUNT(*) AS total,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 1) AS `jan`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 2) AS `fev`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 3) AS `mar`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 4) AS `abr`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 5) AS `mai`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 6) AS `jun`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 7) AS `jul`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 8) AS `ago`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 9) AS `et`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 10) AS `out`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 11) AS `nov`,
    #     SUM(EXTRACT(MONTH FROM dt_solicitacao) = 12) AS `dez`
    # FROM
    #     iluminacao_ordemdeservico
    # WHERE
    #     bairro IS NOT NULL
    #     AND status IN ('0', '1', '2')
    #     AND YEAR(dt_solicitacao) = 2023
    # GROUP BY
    #     bairro
    # '''

    # with connection.cursor() as cursor:
    #     cursor.execute(query)
    #     rows = cursor.fetchall()

    # data = []
    # for row in rows:
    #     bairro = row[0]
    #     total = row[1]
    #     meses = row[2:]
    #     os_por_mes = {meses[i]: value for i, value in enumerate(meses)}
    #     data.append({'bairro': bairro, 'mes': os_por_mes, 'total': total})

    # query = """
    #     SELECT (SELECT COUNT(*) FROM iluminacao_ordemdeservico WHERE status!='f' and bairro = main.bairro GROUP BY bairro) as total, bairro, (SELECT COUNT(dt_solicitacao) FROM iluminacao_ordemdeservico GROUP BY bairro), 
    #         CASE
    #             WHEN strftime('%m', dt_solicitacao) = '01' THEN 'Janeiro'
    #             WHEN strftime('%m', dt_solicitacao) = '02' THEN 'Fevereiro'
    #             WHEN strftime('%m', dt_solicitacao) = '03' THEN 'Março'
    #             WHEN strftime('%m', dt_solicitacao) = '04' THEN 'Abril'
    #             WHEN strftime('%m', dt_solicitacao) = '05' THEN 'Maio'
    #             WHEN strftime('%m', dt_solicitacao) = '06' THEN 'Junho'
    #             WHEN strftime('%m', dt_solicitacao) = '07' THEN 'Julho'
    #             WHEN strftime('%m', dt_solicitacao) = '08' THEN 'Agosto'
    #             WHEN strftime('%m', dt_solicitacao) = '09' THEN 'Setembro'
    #             WHEN strftime('%m', dt_solicitacao) = '10' THEN 'Outubro'
    #             WHEN strftime('%m', dt_solicitacao) = '11' THEN 'Novembro'
    #             WHEN strftime('%m', dt_solicitacao) = '12' THEN 'Dezembro'
    #             ELSE NULL
    #     END AS nome_mes
    #     FROM iluminacao_ordemdeservico as main
    #     WHERE status != 'f'
    #     GROUP BY bairro
    #     ORDER BY bairro;
    #     """
    query = """
        SELECT (SELECT COUNT(*) FROM iluminacao_ordemdeservico WHERE status!='f' and bairro = main.bairro GROUP BY bairro) as total,bairro, 
        COUNT(dt_solicitacao), 
        MONTHNAME(dt_solicitacao) AS nome_mes 
        FROM iluminacao_ordemdeservico as main 
        WHERE status != 'f' 
        GROUP BY bairro, MONTH(dt_solicitacao) 
        ORDER BY bairro; 
    """
    with connection.cursor() as cursor:
        cursor.execute(query)
        results = cursor.fetchall()

    # print(results)
    resultados_formatados = []
    for row in results:
        total, bairro, count_dt_solicitacao, nome_mes = row
        resultado_dict = {
            'bairro': bairro,
            'total': total,
            'mes': nome_mes,
            'total_por_mes': count_dt_solicitacao
        }
        resultados_formatados.append(resultado_dict)

    bairros=[]
    for result in resultados_formatados:  
        bairros.append(result['bairro'])
    bairros_ = list(set(bairros))

    resultados_agrupados = []
    for b in bairros_:
        resultados_agrupados.append(
            {'bairro': b}
        )

    for result in resultados_formatados:
        bairro = result['bairro']
        mes = result['mes']
        total_por_mes = result['total_por_mes']
        total = result['total']
        # print(bairro, mes, total_por_mes, total)
        for i in resultados_agrupados:
            # print(i)
            if i['bairro'] == bairro:
                i[mes]=total_por_mes
                i['total']=total
       
                
    #     else:
            # resultados_agrupados.append({'bairro': bairro, f'{mes}':  total_por_mes, 'total': total})   
    #       for item in resultados_agrupados:
    #           if item.bairro ==  bairro:
    #               item[result.mes]=result.total_por_mes


    # print(resultados_agrupados)
    context = {
        'titulo': apps.get_app_config('iluminacao').verbose_name,   
        'results': resultados_agrupados  
    }
    return render(request, 'iluminacao/painel.html', context)

@login_required
@group_required('os_acesso')
def os_index(request):
    if request.user.is_superuser:
        queryset = OrdemDeServico.objects.all().exclude(status='f')
    else:
        queryset = OrdemDeServico.objects.filter(atendente=request.user ).exclude(status='f')

    if request.method == 'POST':
        # print(request.POST)
        # Obtenha os parâmetros da consulta do formulário
        protocolo = request.POST.get('protocolo')
        tipo_os = request.POST.get('tipo')
        prioridade = request.POST.get('prioridade')
        status = request.POST.get('status')
        bairro = request.POST.get('bairro')
        rua = request.POST.get('rua')
        motivo = request.POST.get('motivo')
        dt_solicitacao1 = request.POST.get('dt_solicitacao1')
        dt_solicitacao2 = request.POST.get('dt_solicitacao2')
        dt_execucao1 = request.POST.get('dt_execucao1')
        dt_execucao2 = request.POST.get('dt_execucao2')
        dt_alteracao1 = request.POST.get('dt_alteracao1')
        dt_alteracao2 = request.POST.get('dt_alteracao2')
        # print(tipo_os)
        # Armazene os parâmetros da consulta na sessão
        request.session['protocolo'] = protocolo
        request.session['tipo_os'] = tipo_os
        request.session['prioridade'] = prioridade
        request.session['status'] = status
        request.session['bairro'] = bairro
        request.session['rua'] = rua
        request.session['motivo'] = motivo
        request.session['dt_solicitacao1'] = dt_solicitacao1
        request.session['dt_solicitacao2'] = dt_solicitacao2
        request.session['dt_execucao1'] = dt_execucao1
        request.session['dt_execucao2'] = dt_execucao2
        request.session['dt_alteracao1'] = dt_alteracao1
        request.session['dt_alteracao2'] = dt_alteracao2
        # print(request.session['tipo_os'])
    else:
        # Recupere os parâmetros da consulta da sessão
        try:
            protocolo = request.session.get('protocolo', '')
            tipo_os = request.session.get('tipo_os', '')
            prioridade = request.session.get('prioridade', '')
            status = request.session.get('status', '')
            bairro = request.session.get('bairro', '')
            rua = request.session.get('rua', '')
            motivo = request.session.get('motivo', '')
            dt_solicitacao1 = request.session.get('dt_solicitacao1', '')
            dt_solicitacao2 = request.session.get('dt_solicitacao2', '')
            dt_execucao1 = request.session.get('dt_execucao1', '')
            dt_execucao2 = request.session.get('dt_execucao2', '')
            dt_alteracao1 = request.session.get('dt_alteracao1', '')
            dt_alteracao2 = request.session.get('dt_alteracao2', '')
        except Exception as E:
            print(E)
            pass
    try:
        # Construa a consulta personalizada
        sql = "SELECT * FROM iluminacao_ordemdeservico WHERE status != 'f'"

        if protocolo:
            sql += f" AND numero LIKE '%{protocolo}%'"
        if tipo_os and tipo_os != 'todos':
            sql += f" AND tipo_id = {tipo_os}"
        if prioridade and prioridade != 'todos':
            sql += f" AND prioridade = '{prioridade}'"
        if status and status != 'todos':
            sql += f" AND status = '{status}'"
        if bairro:
            sql += f" AND bairro LIKE '%{bairro}%'"
        if rua:
            sql += f" AND logradouro LIKE '%{rua}%'"
        if motivo:
            sql += f" AND motivo_reclamacao LIKE '%{motivo}%'"
        if dt_solicitacao1 and dt_solicitacao2:
            sql += f" AND dt_solicitacao BETWEEN '{dt_solicitacao1}' AND '{dt_solicitacao2}'"
        if dt_execucao1 and dt_execucao2:
            sql += f" AND dt_execucao BETWEEN '{dt_execucao1}' AND '{dt_execucao2}'"
        if dt_alteracao1 and dt_alteracao2:
            sql += f" AND dt_alteracao BETWEEN '{dt_alteracao1}' AND '{dt_alteracao2}'"
        # sql += " ORDER BY dt_alteracao, dt_solicitacao"
        # Executar a consulta SQL personalizada
        with connection.cursor() as cursor:
            cursor.execute(sql)
            results = cursor.fetchall()

        # Processar os resultados
        queryset = []
        for row in results:
            # Mapear os campos do modelo e seus valores correspondentes
            data = {
                'id': row[0],
                'numero': row[1],
                'prioridade': row[2],
                'dt_solicitacao': row[3],
                'logradouro': row[4],
                'bairro': row[5],
                'referencia': row[6],
                'motivo_reclamacao': row[7],
                'status': row[8],
                'dt_conclusao': row[9],
                'atendente_id': row[10],
                'cadastrado_por_id': row[11],
                'tipo_id': row[12],
                'pontos_atendidos': row[13],
                'nome_do_contribuinte': row[14],
                'telefone_do_contribuinte': row[15],
                'dt_alteracao': row[16],
                'dt_execucao': row[17],
                'observacao_pontos': row[18]
                # Mapear os demais campos conforme necessário
            }
            ordem_de_servico = OrdemDeServico(**data)
            queryset.append(ordem_de_servico)
            queryset = sorted(queryset, key=lambda x: (x.dt_alteracao if x.dt_alteracao else x.dt_solicitacao), reverse=True)
    except Exception as e:
        print(e)
        pass
    paginator = Paginator(queryset, 30)
    page = request.GET.get('page', 1)
    ordens_de_servico = paginator.get_page(page)

    context = {
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'ordens_de_servico': ordens_de_servico,
        'tipo_os': Tipo_OS.objects.all(),
        'protocolo': protocolo,
        'tipo_os_selecionado': tipo_os,
        'prioridade': prioridade,
        'status': status,
        'bairro': bairro,
        'rua': rua,
        'motivo': motivo,
        'dt_solicitacao1': dt_solicitacao1,
        'dt_solicitacao2': dt_solicitacao2,
        'dt_execucao1': dt_execucao1,
        'dt_execucao2': dt_execucao2,
        'dt_alteracao1': dt_alteracao1,
        'dt_alteracao2': dt_alteracao2,
    }

    return render(request, 'iluminacao/index.html', context)



@login_required
@group_required('os_acesso')
def os_finalizados(request):
    if request.user.is_superuser:
        queryset = OrdemDeServico.objects.filter(status='f')
    else:
        queryset = OrdemDeServico.objects.filter(atendente=request.user, status='f')

    if request.method == 'POST':
        # Obtenha os parâmetros da consulta do formulário
        protocolo = request.POST.get('protocolo')
        tipo_os = request.POST.get('tipo')
        prioridade = request.POST.get('prioridade')
        status = request.POST.get('status')
        bairro = request.POST.get('bairro')
        rua = request.POST.get('rua')
        motivo = request.POST.get('motivo')
        dt_solicitacao1 = request.POST.get('dt_solicitacao1')
        dt_solicitacao2 = request.POST.get('dt_solicitacao2')
        dt_execucao1 = request.POST.get('dt_execucao1')
        dt_execucao2 = request.POST.get('dt_execucao2')
        dt_alteracao1 = request.POST.get('dt_alteracao1')
        dt_alteracao2 = request.POST.get('dt_alteracao2')

        # Armazene os parâmetros da consulta na sessão
        request.session['protocolo_f'] = protocolo
        request.session['tipo_os_f'] = tipo_os
        request.session['prioridade_f'] = prioridade
        request.session['status_f'] = status
        request.session['bairro_f'] = bairro
        request.session['rua_f'] = rua
        request.session['motivo_f'] = motivo
        request.session['dt_solicitacao1_f'] = dt_solicitacao1
        request.session['dt_solicitacao2_f'] = dt_solicitacao2
        request.session['dt_execucao1_f'] = dt_execucao1
        request.session['dt_execucao2_f'] = dt_execucao2
        request.session['dt_alteracao1_f'] = dt_alteracao1
        request.session['dt_alteracao2_f'] = dt_alteracao2
    else:
        # Recupere os parâmetros da consulta da sessão
        protocolo = request.session.get('protocolo_f', '')
        tipo_os = request.session.get('tipo_os_f', '')
        prioridade = request.session.get('prioridade_f', '')
        status = request.session.get('status_f', '')
        bairro = request.session.get('bairro_f', '')
        rua = request.session.get('rua_f', '')
        motivo = request.session.get('motivo_f', '')
        dt_solicitacao1 = request.session.get('dt_solicitacao1_f', '')
        dt_solicitacao2 = request.session.get('dt_solicitacao2_f', '')
        dt_execucao1 = request.session.get('dt_execucao1_f', '')
        dt_execucao2 = request.session.get('dt_execucao2_f', '')
        dt_alteracao1 = request.session.get('dt_alteracao1_f', '')
        dt_alteracao2 = request.session.get('dt_alteracao2_f', '')
    try:
        # Construa a consulta personalizada
        sql = "SELECT * FROM iluminacao_ordemdeservico WHERE status = 'f'"

        if protocolo:
            sql += f" AND numero LIKE '%{protocolo}%'"
        if tipo_os != 'todos':
            sql += f" AND tipo_id = {tipo_os}"
        if prioridade != 'todos':
            sql += f" AND prioridade = '{prioridade}'"
        if status != 'todos':
            sql += f" AND status = '{status}'"
        if bairro:
            sql += f" AND bairro LIKE '%{bairro}%'"
        if rua:
            sql += f" AND logradouro LIKE '%{rua}%'"
        if motivo:
            sql += f" AND motivo_reclamacao LIKE '%{motivo}%'"
        if dt_solicitacao1 and dt_solicitacao2:
            sql += f" AND dt_solicitacao BETWEEN '{dt_solicitacao1}' AND '{dt_solicitacao2}'"
        if dt_execucao1 and dt_execucao2:
            sql += f" AND dt_execucao BETWEEN '{dt_execucao1}' AND '{dt_execucao2}'"
        if dt_alteracao1 and dt_alteracao2:
            sql += f" AND dt_alteracao BETWEEN '{dt_alteracao1}' AND '{dt_alteracao2}'"

        # Executar a consulta SQL personalizada
        with connection.cursor() as cursor:
            cursor.execute(sql)
            results = cursor.fetchall()

        # Processar os resultados
        queryset = []
        for row in results:
            # Mapear os campos do modelo e seus valores correspondentes
            data = {
                'id': row[0],
                'numero': row[1],
                'prioridade': row[2],
                'dt_solicitacao': row[3],
                'logradouro': row[4],
                'bairro': row[5],
                'referencia': row[6],
                'motivo_reclamacao': row[7],
                'status': row[8],
                'dt_conclusao': row[9],
                'atendente_id': row[10],
                'cadastrado_por_id': row[11],
                'tipo_id': row[12],
                'pontos_atendidos': row[13],
                'nome_do_contribuinte': row[14],
                'telefone_do_contribuinte': row[15],
                'dt_alteracao': row[16],
                'dt_execucao': row[17],
                'observacao_pontos': row[18]
                # Mapear os demais campos conforme necessário
            }
            ordem_de_servico = OrdemDeServico(**data)
            queryset.append(ordem_de_servico)
    except:
        pass
    paginator = Paginator(queryset, 30)
    page = request.GET.get('page', 1)
    ordens_de_servico = paginator.get_page(page)

    context = {
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'ordens_de_servico': ordens_de_servico,
        'tipo_os': Tipo_OS.objects.all(),
        'protocolo': protocolo,
        'tipo_os_selecionado': tipo_os,
        'prioridade': prioridade,
        'status': status,
        'bairro': bairro,
        'motivo': motivo,
        'dt_solicitacao1': dt_solicitacao1,
        'dt_solicitacao2': dt_solicitacao2,
        'dt_execucao1': dt_execucao1,
        'dt_execucao2': dt_execucao2,
        'dt_alteracao1': dt_alteracao1,
        'dt_alteracao2': dt_alteracao2,
    }

    return render(request, 'iluminacao/finalizados.html', context)

@login_required
def add_os(request):
    
    form = OS_Form(initial={'tipo': Tipo_OS.objects.get(sigla='IP').id})

    if request.method=='POST':
        form=OS_Form(request.POST)
        if form.is_valid():
            os=form.save(commit=False)
            os.cadastrado_por=Pessoa.objects.get(user=request.user)            
            os.save()
            os.dt_alteracao=os.dt_solicitacao
            os.save()

            return redirect('iluminacao:os_index')                

    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'form': form,
    }

    return render(request, 'iluminacao/adicionar_os.html', context)

@login_required
@group_required('os_acesso')
def detalhes_os(request, id):
    pessoa = Pessoa.objects.get(user=request.user)
    os = OrdemDeServico.objects.get(id=id)
    form_mensagem = NovaMensagemForm(initial={'os': os.id, 'pessoa': pessoa.id})
    try:
        os_ext=OS_ext.objects.get(os=os)        
    except:
        os_ext = None         
    if request.method=='POST': 
        form_mensagem=NovaMensagemForm(request.POST, request.FILES)
        if form_mensagem.is_valid():
           msg=form_mensagem.save(commit=False)
           msg.os=os
           msg.pessoa=pessoa
           msg.save()
           form_mensagem = NovaMensagemForm(initial={'os': os.id, 'pessoa': pessoa.id})
           os.message_status='2'
           os.save()
    else:
        if os.message_status=='2':
            os.message_status='1'
            os.save()

    linha_tempo=OS_Linha_Tempo.objects.filter(os=os)
    context={
        'form_mensagem': form_mensagem,
        'linha_tempo': linha_tempo,
        'STATUS': STATUS_CHOICES,
        'PRIORIDADES': PRIORIDADE_CHOICES, 
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'os': os,
        'os_ext': os_ext
    }
    return render(request, 'iluminacao/detalhes_os.html', context)

@login_required
@group_required('os_acesso')
def change_status_os(request, id, opcao):
    os=OrdemDeServico.objects.get(id=id)
    os.status=opcao
    print(opcao)
    if opcao=='f':
        os.dt_conclusao=datetime.now()
        os.finalizado_por=request.user
    os.save()
    return redirect('iluminacao:detalhes_os', id=id)

@login_required
@group_required('os_acesso')
def change_prioridade_os(request, id, opcao):
    os=OrdemDeServico.objects.get(id=id)
    os.prioridade=opcao
    os.save()
    return redirect('iluminacao:detalhes_os', id=id)

@login_required
@group_required('os_acesso')
def atender_os(request, id):
    os=OrdemDeServico.objects.get(id=id)
    os.atendente=request.user
    os.save()
    return redirect('iluminacao:detalhes_os', id=id)

@login_required
@group_required('os_acesso')
def funcionarios_listar(request):
    funcionarios=Funcionario_OS.objects.all()
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'funcionarios': funcionarios
    }
    return render(request, 'equipe/funcionarios.html', context)

@login_required
@group_required('os_acesso')
def funcionario_cadastrar(request):
    if request.method=='POST':
        form=Funcionario_Form({'pessoa':request.POST['pessoa'], 'nivel': request.POST['nivel'], 'tipo_os': [1]})
        if form.is_valid():
            form.save()
            funcionario=Funcionario_OS()
            return redirect('iluminacao:funcionarios')
        # else:
            # print(form.errors)
    else:
        form=Funcionario_Form(initial={'tipo_os': Tipo_OS.objects.get(sigla='IP')})
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'form': form
    }
    return render(request, 'equipe/funcionarios_cadastrar.html', context)

@login_required
@group_required('os_acesso')
def funcionario_editar(request, id):
    funcionario=Funcionario_OS.objects.get(id=id)
    form=Funcionario_Form_editar(instance=funcionario)
    if request.method=='POST':
        form=Funcionario_Form_editar(request.POST, instance=funcionario)
        if form.is_valid():
            form.save()
            return redirect('funcionarios')
    context={
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'form': form,
        'funcionario': funcionario
    }     
    return render(request, 'equipe/funcionarios_editar.html', context)

@login_required
@group_required('os_acesso')
def funcionario_deletar(request, id):
    funcionario=Funcionario_OS.objects.get(id=id)
    funcionario.delete()

    return redirect('iluminacao:funcionarios')

@login_required
@group_required('os_acesso')
def atribuir_equipe(request, id):
    try:
        instancia=OS_ext.objects.get(os=id)
        form=Equipe_Form(instance=instancia)        
    except Exception as e:
        form=Equipe_Form(initial={'os': id})
        instancia=None
        
    if request.method=='POST':
        if instancia:
            form=Equipe_Form(request.POST, instance=instancia)
        else:
            form=Equipe_Form(request.POST)
        if form.is_valid:
            form.save()
            return redirect('iluminacao:detalhes_os', id)
    context={
            'titulo': apps.get_app_config('iluminacao').verbose_name,   
            'form':form,
        }
    return render(request, 'iluminacao/adicionar_ext.html', context)

@login_required
@group_required('os_acesso')
def pontos_os(request, id):
    instancia=OrdemDeServico.objects.get(id=id)
    # print(instancia.dt_execucao)
    form=OS_Form_Ponto(instance=instancia)            
        
    if request.method=='POST':       
        form=OS_Form_Ponto(request.POST, instance=instancia)
        if form.is_valid:
            form.save()
            return redirect('iluminacao:detalhes_os', id)
    else:
        # Converta a data para o formato correto AAAA-MM-DD antes de inicializar o formulário
        if instancia.dt_execucao is not None:
            # Convert the date to the correct format (YYYY-MM-DD) before initializing the form
            instancia.dt_execucao = instancia.dt_execucao.strftime('%Y-%m-%d')
        form = OS_Form_Ponto(instance=instancia)
    context={
            'titulo': apps.get_app_config('iluminacao').verbose_name,   
            'form':form,
        }
    return render(request, 'iluminacao/adicionar_ext.html', context)

from django.http import HttpResponse
from django.template.loader import render_to_string
# from weasyprint import HTML
from django.utils.dateparse import parse_date
from django.middleware.csrf import get_token

@login_required
@group_required('os_acesso')
def imprimir_os(request, id):
    lista_de_os=[OrdemDeServico.objects.get(id=id)]
    context={
        'lista_de_os': lista_de_os
    }
    return render(request, 'iluminacao/imprimir_os.html', context)

@login_required
@group_required('os_acesso')
def imprimir_varias_os(request, ids):
    ids=ids.split('-')
    ids.pop()
    lista_de_os=[]
    for i in ids:
        lista_de_os.append(OrdemDeServico.objects.get(id=i))
    # print(len(lista_de_os))
    # print(ids)
    context={
        'lista_de_os': lista_de_os
    }
    return render(request, 'iluminacao/imprimir_os.html', context)

@login_required
@group_required('os_acesso')
def imprimir_todas_os(request):
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')

        try:
            if start_date and end_date:
                start_date = parse_date(start_date)
                end_date = parse_date(end_date)
                if start_date is None or end_date is None:
                    raise ValueError("Invalid date format")
                lista_de_os = OrdemDeServico.objects.filter(dt_solicitacao__range=[start_date, end_date])
            else:
                lista_de_os = OrdemDeServico.objects.all()
        except ValueError:
            return HttpResponse("Data inválida fornecida", status=400)
        
        context = {'lista_de_os': lista_de_os}    
        template = 'iluminacao/imprimir_todas_os.html'
        return render(request, template, context)
    
    csrf_token = get_token(request)
    html_form = f'''
        <html>
        <head><title>Filtrar Ordens de Serviço</title></head>
        <body>
            <h1>Filtrar Ordens de Serviço por Data</h1>
            <form method="post">
                <input type="hidden" name="csrfmiddlewaretoken" value="{csrf_token}">
                <label for="start_date">Data Inicial:</label>
                <input type="date" id="start_date" name="start_date" required>
                <label for="end_date">Data Final:</label>
                <input type="date" id="end_date" name="end_date" required>
                <button type="submit">Gerar PDF</button>
            </form>
        </body>
        </html>
        '''
    return HttpResponse(html_form)

@login_required
@group_required('os_acesso')
def graficos(request):
    
    pontos_por_bairro_finalizados = OrdemDeServico.objects.filter(status='f').values('bairro').annotate(total=Sum('pontos_atendidos')).order_by('-total')
    pontos_por_bairro_nao_finalizados = OrdemDeServico.objects.exclude(status='f').values('bairro').annotate(total=Sum('pontos_atendidos')).order_by('-total')
    
    # Consulta para os bairros com ordens de serviço finalizadas
    servicos_por_bairro_finalizados = OrdemDeServico.objects.filter(status='f').values('bairro').annotate(total=Count('id')).order_by('-total')
    # print(len(servicos_por_bairro_finalizados))
    # Consulta para os bairros com ordens de serviço não finalizadas
    servicos_por_bairro_nao_finalizados = OrdemDeServico.objects.exclude(status='f').values('bairro').annotate(total=Count('id')).order_by('-total')
    # print(len(servicos_por_bairro_nao_finalizados))
    # Combina os resultados em uma única lista, se necessário
    servicos_por_bairro = pontos_por_bairro_finalizados.union(pontos_por_bairro_nao_finalizados)

    finalizados = OrdemDeServico.objects.filter(status='f').count()
    nao_finalizados = OrdemDeServico.objects.exclude(status='f').count()

    # Quantitativo de OS finalizadas por tipo
    os_finalizadas_por_tipo = OrdemDeServico.objects.filter(status='f').values('tipo__sigla').annotate(
        total_finalizadas=Count('id'),
    ).order_by('-total_finalizadas')

    # Quantitativo de OS não finalizadas por tipo
    os_nao_finalizadas_por_tipo = OrdemDeServico.objects.exclude(status='f').values('tipo__sigla').annotate(
        total_nao_finalizadas=Count('id'),
    ).order_by('-total_nao_finalizadas')

    os_por_funcionario = Funcionario_OS.objects.annotate(total_os=Count('os_ext__os')).order_by('-total_os')
    pontos_por_funcionario = Funcionario_OS.objects.annotate(total_pontos=Sum('os_ext__os__pontos_atendidos')).order_by('-total_pontos')

    today = timezone.now()
    last_12_months = today - timedelta(days=365)

    # Consulta SQL bruta para obter os dados
    sql_query = """
        SELECT
            DATE_FORMAT(dt_solicitacao, '%%b %%Y') AS month_label,
            SUM(CASE WHEN status = 'f' THEN 1 ELSE 0 END) AS total_finalizadas,
            SUM(CASE WHEN status <> 'f' THEN 1 ELSE 0 END) AS total_nao_finalizadas
        FROM
            iluminacao_ordemdeservico
        WHERE
            dt_solicitacao >= %s
        GROUP BY
            month_label
        ORDER BY
            dt_solicitacao
    """

    with connection.cursor() as cursor:
        cursor.execute(sql_query, [last_12_months])
        result = cursor.fetchall()

    # Extraindo dados para o gráfico
    servicos_grafico_linha_dados = []

    for row in result:
        mes, total_finalizadas, total_nao_finalizadas = row

        # Anexar dados às respectivas listas
       
        servicos_grafico_linha_dados.append([month_translate(mes), int(total_finalizadas), int(total_nao_finalizadas)])
    context = {
        'pontos_por_bairro_nao_finalizados': pontos_por_bairro_nao_finalizados,
        'pontos_por_bairro_finalizados': pontos_por_bairro_finalizados,
        'servicos_por_bairro_finalizados': servicos_por_bairro_finalizados,
        'servicos_por_bairro_nao_finalizados': servicos_por_bairro_nao_finalizados,
        'finalizados': finalizados,
        'nao_finalizados': nao_finalizados,
        'os_por_funcionario': os_por_funcionario,
        'pontos_por_funcionario': pontos_por_funcionario,
        'os_finalizadas_por_tipo': os_finalizadas_por_tipo,
        'os_nao_finalizadas_por_tipo': os_nao_finalizadas_por_tipo,
        'titulo': apps.get_app_config('iluminacao').verbose_name,
        'servicos_grafico_linha_dados': servicos_grafico_linha_dados
    }
    return render(request, 'iluminacao/graficos.html', context)


@login_required
@group_required('os_acesso')
def graficos_ver_mais(request, tipo, subtipo):
    context = {
        'tipo': tipo,
        'titulo': apps.get_app_config('iluminacao').verbose_name,
    }
    workbook = Workbook()
    if tipo == 'pontos-por-bairro':
        pontos_por_bairro = OrdemDeServico.objects.values('bairro').annotate(total=Sum('pontos_atendidos')).order_by('-total')
        dados = [{'y': item['bairro'], 'total': item['total']} for item in pontos_por_bairro]
        context['dados'] = dados
        context['y'] = 'Bairros'
        context['x'] = 'Pontos'
        
        # Crie uma planilha no workbook
        planilha = workbook.active
        planilha.title = 'Pontos por Bairro'
        # Adicione os cabeçalhos das colunas
        planilha['A1'] = 'Bairros'
        planilha['B1'] = 'Pontos'
        # Preencha os dados
        for index, item in enumerate(dados, start=2):
            planilha.cell(row=index, column=1, value=item['y'])
            planilha.cell(row=index, column=2, value=item['total'])

    elif tipo == 'os-por-bairro':
        os_por_bairro = OrdemDeServico.objects.values('bairro').annotate(total=Count('id')).order_by('-total')
        dados = [{'y': item['bairro'], 'total': item['total']} for item in os_por_bairro]
        context['dados'] = dados
        context['y'] = 'Bairros'
        context['x'] = 'Nº de OS'
        
        planilha = workbook.active
        planilha.title = 'OS por Bairro'
        # Adicione os cabeçalhos das colunas
        planilha['A1'] = 'Bairros'
        planilha['B1'] = 'Nº de OS'
        # Preencha os dados
        for index, item in enumerate(dados, start=2):
            planilha.cell(row=index, column=1, value=item['y'])
            planilha.cell(row=index, column=2, value=item['total'])

    elif tipo == 'pontos-por-funcionario':
        pontos_por_funcionario = Funcionario_OS.objects.annotate(total=Sum('os_ext__os__pontos_atendidos')).order_by('-total')
        dados = [{'y': str(item), 'total': item.total} for item in pontos_por_funcionario]
        context['dados'] = dados
        context['y'] = 'Funcionarios'
        context['x'] = 'Pontos'

         # Crie uma planilha no workbook
        planilha = workbook.active
        planilha.title = 'Pontos por Funcionário'
        # Adicione os cabeçalhos das colunas
        planilha['A1'] = 'Funcionários'
        planilha['B1'] = 'Pontos'
        # Preencha os dados
        for index, item in enumerate(dados, start=2):
            planilha.cell(row=index, column=1, value=item['y'])
            planilha.cell(row=index, column=2, value=item['total'])

    else:
        return redirect('iluminacao:kpi')
    if subtipo=='imprimir':
        return render(request, 'iluminacao/graficos_ver_mais_imprimir.html', context)
    elif subtipo=='download':
        data_atual = datetime.now()
        data_formatada = data_atual.strftime("%d-%m-%y")
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename={tipo}-{data_formatada}.xlsx'        
        workbook.save(response)
        return response
    return render(request, 'iluminacao/graficos_ver_mais.html', context)

@login_required
@group_required('os_acesso')
def mudadados(request):
    finalizados = OrdemDeServico.objects.filter(status='f')
    count = 0
    for item in finalizados:
        mensagens = OS_Linha_Tempo.objects.filter(os=item)
        for mensagem in mensagens:
            if mensagem.mensagem[0] == '*':
                data = mensagem.mensagem[1:11]
                dt = datetime.strptime(data, "%d/%m/%Y")
                dt = dt.strftime("%Y-%m-%d")
                item.dt_conclusao = dt
                item.save()
                msg = mensagem.mensagem.replace(mensagem.mensagem[0:11], "")
                if msg == "":
                   msg = "Data de conclusão alterada devido a registro retroativo."
                mensagem.mensagem = msg
                mensagem.save()
    return render(request, 'template.html')

def alterar_equipes(request):
    os_finalizadas = OrdemDeServico.objects.filter(status='f')
    for os in os_finalizadas:
        extensoes_dos_OS = OS_ext.objects.filter(os=os)
        for extensao in extensoes_dos_OS:
            if extensao.cod_veiculo != '':
                codigo_do_veiculo = extensao.cod_veiculo.split('/')
                for funcionario in codigo_do_veiculo:
                    funcionario_pessoa = Funcionario_OS.objects.get(funcionario)
                    # print(funcionario_pessoa)
                    #funcionarios = Funcionario_OS.objects.filter(nome__startswith=funcionario)
                    #print(funcionarios)

    return render(request, 'template.html')
    

@login_required
@group_required('os_acesso')
def salvar_contagem_os(request):
    # Limpa as tabelas existentes antes de salvar os novos dados
    TotalOSPorSemanaAno.objects.all().delete()
    TotalOSPorMesAno.objects.all().delete()

    # Chama os métodos estáticos para obter e salvar os dados
    OrdemDeServico.total_os_por_semana_ano()
    OrdemDeServico.total_os_por_mes_ano()

    return redirect('iluminacao:contagem_os')

@login_required
@group_required('os_acesso')
def contagem_os(request):
    # Obtém a contagem de OS por semana e ano
    total_os_semana_ano = TotalOSPorSemanaAno.objects.all()

    # Obtém a contagem de OS por mês
    total_os_mes_ano = TotalOSPorMesAno.objects.all()

    context = {
        'total_os_semana_ano': total_os_semana_ano,
        'total_os_mes': total_os_mes_ano
    }

    return render(request, 'iluminacao/contagem_os.html', context)



