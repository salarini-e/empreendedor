from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db import IntegrityError
from django.contrib import messages
from django.http import Http404
from autenticacao.functions import aluno_required
from .models import *
from .forms import *
from datetime import date, datetime
from .models import *
from .forms import *
from autenticacao.forms import Form_Pessoa, Form_Alterar_Pessoa
# from eventos.models import Evento
from django.apps import apps
from random import shuffle
from django.urls import reverse
from django.core.paginator import Paginator
from django.contrib.auth.decorators import user_passes_test
from django.http import JsonResponse
from django.utils import timezone
def index(request):
    # try:
    #     eventos = Evento.objects.filter(app_name='cursos', is_destaque = True)
    # except:
    eventos=[]
    
    cursos = list(Curso.objects.filter(tipo='C', ativo=True).order_by('?')[:9])
    palestras = list(Curso.objects.filter(tipo='P', ativo=True).order_by('?')[:4])
    shuffle(cursos)
    context = { 
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name,
        'eventos': eventos,
        'cursos': cursos,   
        'cursos_en': Curso_Ensino_Superior.objects.all()[:4],
        'palestras': palestras 
    }

    return render(request, 'cursos_empresariais/index.html', context)

def retirar_duplicadas(request):
    alertas = Alertar_Aluno_Sobre_Nova_Turma.objects.all()
    count = 0
    for alerta in alertas:
        if Alertar_Aluno_Sobre_Nova_Turma.objects.filter(aluno=alerta.aluno, curso=alerta.curso, alertado=False).count() > 1:
            alerta.delete()
            count+=1
    return HttpResponse(f'{count} duplicadas removidas')

def cursos(request):
    tipo='cursos'
    form = Aluno_form()
    categorias = Categoria.objects.all()
    cursos = []
    if tipo == 'cursos':        
            cursos=Curso.objects.filter(tipo='C', ativo=True)
    elif tipo == 'palestras':
            cursos=Curso.objects.filter(tipo='P', ativo=True)    

    context = {
        'categorias':categorias,
        'cursos': cursos,
        'form': form,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name,        
        'tipo': tipo,
    }
    if request.user.is_authenticated:
        pessoa=Pessoa.objects.get(user=request.user)
        try:
            aluno=Aluno.objects.get(pessoa=pessoa)
            context['alertas']=Alertar_Aluno_Sobre_Nova_Turma.objects.filter(aluno=aluno, alertado=False)
        except:
            pass
        
    if tipo == 'cursos':
        return render(request, 'cursos_empresariais/cursos.html', context)
    elif tipo == 'palestras':  
        return render(request, 'cursos_empresariais/palestras.html', context)
    raise Http404("Página não encontrada")  

def cursos_filtrado(request, tipo, filtro):
    tipo='curso'
    form = Aluno_form()
    categorias = Categoria.objects.all()
    cursos = []
    if tipo == 'cursos':       
       cursos=Curso.objects.filter(tipo='C', categoria__nome=filtro, ativo=True)
    elif tipo == 'palestras':                
        cursos=Curso.objects.filter(tipo='P', categoria__nome=filtro, ativo=True)

    context = {
        'categorias': categorias,
        'cursos': cursos,
        'form': form,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name + ' - ' + tipo.capitalize(), 
        'filtro': filtro,
        'tipo': tipo
    }
    if tipo == 'cursos':
        return render(request, 'cursos_empresariais/cursos.html', context)
    elif tipo == 'palestras':  
        return render(request, 'cursos_empresariais/palestras.html', context)
    raise Http404("Página não encontrada")

def curso_detalhe(request, tipo, id):    
    curso=Curso.objects.get(id=id)    
    context={
        'curso': curso,
        'tipo': tipo,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Detalhes',
        'turmas': Turma.objects.filter(curso=curso)
    }
    return render(request, 'cursos_empresariais/curso_detalhe.html', context)

@login_required
def candidatar(request, id):

    curso = Curso.objects.get(id=id)
    form = Aluno_form(initial={'curso': curso})
    if request.method == 'POST':
        form = Aluno_form(request.POST)
        if form.is_valid():
            form.save()

    context = {
        'form': form,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name + ' - Inscrição - ' + curso.nome,
    }
    return render(request, 'cursos_empresariais/cadastrar_candidato.html', context)


def prematricula(request):
    form = Aluno_form(prefix="candidato")
    form_responsavel = CadastroResponsavelForm(prefix="responsavel")

    categorias = Categoria.objects.all()
    cursos = []

    for i in categorias:
        cursos.append(
            {'categoria': i, 'curso': Curso.objects.filter(categoria=i, ativo=True)})

    if request.method == 'POST':
        dtnascimento_cp = request.POST.get("candidato-dt_nascimento")
        form = Aluno_form(request.POST, prefix="candidato")
        form_responsavel = CadastroResponsavelForm(
            request.POST, prefix="responsavel")

        try:
            dtnascimento_hr = datetime.strptime(dtnascimento_cp, "%d-%m-%Y")
        except:
            dtnascimento_hr = datetime.strptime(dtnascimento_cp, "%Y-%m-%d")

        dt_nascimento = dtnascimento_hr.date()

        today = date.today()
        age = today.year - dt_nascimento.year - \
            ((today.month, today.day) < (dt_nascimento.month, dt_nascimento.day))
        teste = True
        candidato = False

        try:
            cpf = request.POST['cpf']
            candidato = Aluno.objects.get(cpf=cpf)
        except Exception as e:
            pass

        if candidato:        
            turma = Turma.objects.get(id=i)
            if candidato:
                try:
                    Matricula.objects.get(
                        candidato=candidato, turma__curso=turma.curso)
                    messages.error(
                        request, 'Candidato já matriculado no curso ' + turma.curso.nome)
                    return redirect('cursos_empresariais:curso_detalhe')
                except:
                    pass

            if (turma.idade_minima is not None and age < turma.idade_minima) or (turma.idade_maxima is not None and age > turma.idade_maxima):
                teste = False

        if form.is_valid() and teste:
            candidato = form.save(commit=False)

            if age < 18:

                if form_responsavel.is_valid():

                    responsavel = form_responsavel.save(commit=False)
                    responsavel.aluno = candidato

                else:
                    return redirect('cursos_empresariais:curso_detalhe')

                responsavel.save()
                candidato.save()

            else:

                candidato.save()

            for i in request.POST.getlist('turmas'):
                Matricula.objects.create(
                    aluno=candidato, turma=turma, status='c')

            messages.success(
                request, 'Pré-inscrição realizada com sucesso! Aguarde nosso contato para finalizar inscrição.')
            return redirect('/')
        else:
            if not teste:
                messages.error(
                    request, 'Não foi possível realizar a inscrição na turma: A idade não atende a faixa etária da turma.')
                return redirect('/prematricula')

    context = {
        'form': form,
        'form_responsavel': form_responsavel,
        'categorias': cursos,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name + ' - 001',
    }
    return render(request, 'cursos_empresariais/pre_matricula.html', context)

def alterarCad(request):
    return render(request, 'cursos_empresariais/alterar_cad.html')


def resultado(request):
    return render(request, 'cursos_empresariais/resultado.html')

@login_required
def matricular(request, tipo, id):
    curso=Curso.objects.get(id=id)
    pessoa=Pessoa.objects.get(user=request.user)

    #checa se já é aluno para por informações no formulario
    try:
        aluno=Aluno.objects.get(pessoa=pessoa)
        form = Aluno_form(prefix="candidato", instance=aluno)
        try:            
            form_responsavel = CadastroResponsavelForm(prefix="responsavel", instance=aluno)
        except:
            pass
    except Exception as E:        
        aluno=None     
        form = Aluno_form(prefix="candidato")    
        form_responsavel = CadastroResponsavelForm(prefix="responsavel")

    # Checa a idade e se precisa de responsavel
    dtnascimento = pessoa.dt_nascimento    
    today = date.today()
    age = today.year - dtnascimento.year - \
            ((today.month, today.day) < (dtnascimento.month, dtnascimento.day))    
    # precisa_responsavel=age<18
    precisa_responsavel=False
    
    if request.method == 'POST':
        
        form = Aluno_form(request.POST, prefix="candidato", instance=aluno)
        if precisa_responsavel:
            form_responsavel = CadastroResponsavelForm(
                request.POST, prefix="responsavel", instance=aluno)

        
        teste = True        
        
        candidato = aluno
        

        #checando idade minima e maxima para o curso
        turmas=Turma.objects.filter(curso=curso)
        pode_cursar=True
        if len(turmas)!=0:
            for turma in turmas:
                if (turma.idade_minima is not None and age < turma.idade_minima) or (turma.idade_maxima is not None and age > turma.idade_maxima):
                    teste = False
                    if not teste:
                        print('entrou aqui')
                        pode_cursar=False
                        teste=True
            
        #validação das informações do forms
        if form.is_valid() and pode_cursar:
            if candidato:
                for turma in turmas:
                    try:
                        Matricula.objects.get(
                            aluno=candidato, turma__curso=turma.curso)
                        messages.error(
                            request, 'Candidato já matriculado no curso: ' + turma.curso.nome)
                        return redirect(reverse('cursos_empresariais:curso_detalhe', args=[tipo, id]))
                    except Exception as E:                        
                        pass
            candidato = form.save(commit=False)

            if precisa_responsavel:
                if form_responsavel.is_valid():
                    responsavel = form_responsavel.save(commit=False)
                    responsavel.aluno = candidato
                else:       
                    messages.warning(
                    request, 'Preencha corretamente os campos do formulário!')
                    context = {
                        'age': age,
                        'form': form,
                        'form_responsavel': form_responsavel,     
                        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Inscrever-se',
                        'curso': curso,
                        'pessoa': pessoa
                    }             
                    return render(request, 'cursos_empresariais/pre_matricula.html', context)
                candidato.pessoa=pessoa
                candidato.save()                                
                responsavel.save()
                
            else:
                candidato.pessoa=pessoa
                # candidato.disponibilidade.clear()
                # for i in request.POST.getlist('candidato-disponibilidade'):
                #     disponibilidade=Disponibilidade.objects.get(id=i)
                #     candidato.disponibilidade.add(disponibilidade)                    
                candidato.save()                    
                # print('ops')
                

            #criando matricula como candidato na turma
            try:
                turma=Turma.objects.filter(curso=curso, status='pre')
                turma_disponibilidade =[]
                # candidato_disponibilidade=[]
                # for i in turma.disponibilidade.all():
                #         turma_disponibilidade.append(i.disponibilidade)
                # for i in candidato.disponibilidade.all():
                #         candidato_disponibilidade.append(i.disponibilidade)
                # pode_entrar = any(disponibilidade in turma_disponibilidade for disponibilidade in candidato_disponibilidade)
                # if pode_entrar:
                Matricula.objects.create(
                        aluno=candidato, turma=turma[0], status='c')
                # else:
                #     messages.error(
                #     request, 'Não foi possível realizar a matricula na turma, pois sua disponibilidade não é compátivel com o da turma aberta.')
                #     context = {
                #             'age': age,
                #             'form': form,
                #             'form_responsavel': form_responsavel,     
                #             'titulo': apps.get_app_config('cursos_empresariais').verbose_name,
                #             'curso': curso,
                #             'pessoa': pessoa
                #         }
                #     return render(request, 'cursos_empresariais/pre_matricula.html', context)
            except:     
                try:
                    turma=Turma.objects.filter(curso=curso, status='acc')
                    # turma_disponibilidade=[]
                    # candidato_disponibilidade=[]
                    # for i in turma.disponibilidade.all():
                    #         turma_disponibilidade.append(i.disponibilidade)
                    # for i in candidato.disponibilidade.all():
                    #         candidato_disponibilidade.append(i.disponibilidade)
                    # pode_entrar = any(disponibilidade in turma_disponibilidade for disponibilidade in candidato_disponibilidade)
                    # if pode_entrar:
                    Matricula.objects.create(
                            aluno=candidato, turma=turma[0], status='c')
                    # else:
                    #     messages.error(
                    #     request, 'Não foi possível realizar a matricula na turma, pois sua disponibilidade não é compátivel com o da turma aberta.')
                    #     context = {
                    #         'age': age,
                    #         'form': form,
                    #         'form_responsavel': form_responsavel,     
                    #         'titulo': apps.get_app_config('cursos_empresariais').verbose_name,
                    #         'curso': curso,
                    #         'pessoa': pessoa
                    #     }
                    #     return render(request, 'cursos_empresariais/pre_matricula.html', context)
                except:
                    # pode_entrar=False
                    if Alertar_Aluno_Sobre_Nova_Turma.check_duplicate(aluno=candidato,curso=curso):
                        messages.error(request, 'Você já realizou uma pré-inscrição para este curso! Aguarde nosso contato para finalizar sua inscrição.')
                    else:
                        Alertar_Aluno_Sobre_Nova_Turma.objects.create(
                            aluno=candidato,
                            curso=curso
                        )
                        messages.success(request, "Você acaba de realizar sua pré-inscrição para o curso! Entraremos em contato por e-mail ou WhatsApp para finalizar sua inscrição nesta edição do curso, ou para agendar sua participação em uma próxima, caso o número de vagas já tenha sido preenchido!")
                    return redirect(reverse('cursos_empresariais:matricula', args=[tipo,id]))
            
            messages.success(
                request, 'Pré-inscrição realizada com sucesso! Aguarde nosso contato para finalizar inscrição.')
            print('passou aqui 2')
            return redirect(reverse('cursos_empresariais:curso_detalhe', args=[tipo, id]))
        else:
            print('passou aqui')
            if not teste:
                messages.error(
                    request, 'Não foi possível realizar a inscrição na turma: A idade não atende a faixa etária da turma.')
                return redirect(reverse('cursos_empresariais:curso_detalhe', args=[tipo, id]))

    context = {
        'age': age,
        'form': form,
        'form_responsavel': form_responsavel,     
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Inscrever-se',
        'curso': curso,
        'pessoa': pessoa
    }
    return render(request, 'cursos_empresariais/pre_matricula.html', context)

from django.http import HttpResponse
from openpyxl import Workbook

def exportar_para_excel(request):
    # Filtrar as turmas com status 'pre'
    cursos = Curso.objects.all().order_by('nome')
    
    # Criar um workbook do Excel
    wb = Workbook()
    # Criar uma planilha
    ws = wb.active
    ws.title = "Alunos Interessados"
    
    # Adicionar cabeçalho à planilha
    ws.append(["Curso", "Aluno", "Envio do pedido", "Telefone", "Email"])

    # Iterar sobre as turmas
    for curso in cursos:
        # Filtrar os alertas para esta turma e após a data de inclusão da turma
        alertas = Alertar_Aluno_Sobre_Nova_Turma.objects.filter(curso=curso, alertado=False)
        # Adicionar os alunos alertados à planilha
        for alerta in alertas:
            ws.append([curso.nome, alerta.aluno.pessoa.nome, alerta.dt_inclusao, alerta.aluno.pessoa.telefone, alerta.aluno.pessoa.email ]) 

    # Definir o nome do arquivo
    dataHora = datetime.datetime.now()
    dataHora_formatada = dataHora.strftime("%Y%m%d_%H%M%S")
    file_name = f"alunos_interessados_{dataHora_formatada}.xlsx"
    
    # Criar uma resposta HTTP para o arquivo Excel
    response = HttpResponse(content_type="application/ms-excel")
    response["Content-Disposition"] = f"attachment; filename={file_name}"
    
    # Salvar o workbook e escrever na resposta HTTP
    wb.save(response)

    return response

def exportar_para_excel_por_turma(request):
    # Filtrar os cursos, excluindo o CEVEST
    # cursos = Curso.objects.exclude(categoria__nome='CEVEST')
    cursos = Curso.objects.all()

    # Criar um novo arquivo Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Alunos por Curso"

    # Adicionar cabeçalhos
    ws.append(['Curso', 'Aluno', 'Email', 'Telefone', 'Status da Matrícula'])

    # Iterar sobre os cursos
    for curso in cursos:
        # Filtrar matrículas por curso
        matriculas = Matricula.objects.filter(turma__curso=curso)

        # Iterar sobre as matrículas
        for matricula in matriculas:
            # Adicionar detalhes do aluno e da matrícula ao Excel
            ws.append([curso.nome, matricula.aluno.pessoa.nome, matricula.aluno.pessoa.email, matricula.aluno.pessoa.telefone, matricula.get_status_display()])

    # Criar a resposta do HTTP com o conteúdo do arquivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=alunos_por_curso.xlsx'
    wb.save(response)

    return response

def ensino_superior(request):
    context = {
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Ensino Superior',
        'cursos': Curso_Ensino_Superior.objects.all()
    }
    return render(request, 'cursos_empresariais/ensino_superior.html', context)

# def ensino_superior_detalhe(request, id):    
#     curso=Curso.objects.get(id=id)
#     context={
#         'curso': curso,
#         'tipo': tipo,
#         'titulo': apps.get_app_config('cursos_empresariais').verbose_name,
#         'turmas': Turma.objects.filter(curso=curso)
#     }
#     return render(request, 'cursos_empresariais/curso_detalhe.html', context)

def ensino_tecnico(request):
    context = {
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Ensino Técnico',
        # 'cursos': Curso_Ensino_Superior.objects.all()
    }
    return render(request, 'cursos_empresariais/ensino_tecnico.html', context)

def curriculo_vitae(request):
    context = {
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Currículo Vitae',
    }
    return render(request, 'cursos_empresariais/curriculo_vitae.html', context)

def area_do_estudante(request):
    pessoa=Pessoa.objects.get(user=request.user)
    try:
        aluno=Aluno.objects.get(pessoa=pessoa)
    except:
        messages.error(request, 'Você não é um aluno! Primeiro você deve se inscrever em um curso.')
        return redirect('cursos_empresariais:cursos', tipo='cursos')
    matriculas=Matricula.objects.filter(aluno=aluno).order_by('-turma__data_inicio')
    alertas=Alertar_Aluno_Sobre_Nova_Turma.objects.filter(aluno=aluno, alertado=False).order_by('-dt_inclusao')
    context={
        'matriculas': matriculas,
        'alertas': alertas,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Área do Estudante'
    }
    return render(request, 'cursos_empresariais/area_do_estudante.html', context)

def editar_cadastro(request):    
    pessoa=Pessoa.objects.get(user=request.user)
    form_pessoa=Form_Alterar_Pessoa(instance=pessoa)    
    if request.method == 'POST':
        form_pessoa=Form_Alterar_Pessoa(request.POST, instance=pessoa)
        if form_pessoa.is_valid:
            form_pessoa.save()
    context={        
        'form_pessoa': form_pessoa,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Editar Cadastro'
    }
    return render(request, 'cursos_empresariais/editar_cadastro.html', context)

def editar_cadastro_pessoa(request):    
    pessoa=Pessoa.objects.get(user=request.user)
    form_pessoa=Form_Pessoa(instance=pessoa)    
    if request.method == 'POST':
        form_pessoa=Form_Pessoa(request.POST, instance=pessoa)
        if form_pessoa.is_valid:
            form_pessoa.save()
            return redirect('cursos_empresariais:home')
    context={        
        'form_pessoa': form_pessoa,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Editar Cadastro'
    }
    return render(request, 'cursos_empresariais/editar_cadastro_pessoa.html', context)

def alterar_senha(request):
    form_user = PasswordChangeCustomForm(user=request.user)

    if request.method == 'POST':
        form_user = PasswordChangeCustomForm(user=request.user, data=request.POST)
        if form_user.is_valid():
            form_user.save()
            return redirect('cursos_empresariais:home')

    context = {        
        'form_user': form_user,
        'titulo': apps.get_app_config('cursos_empresariais').verbose_name+' - Alterar Senha'
    }
    return render(request, 'cursos_empresariais/altera_senha.html', context)


#ADMINISTRATIVO


@staff_member_required
def enviar_email(aluno, turma):
    try:
        subject = f'Inscrição no curso {turma.curso.nome}'
        from_email = settings.EMAIL_HOST_USER
        to = [aluno.email]
        text_content = 'This is an important message.'
        html_content = render_to_string('email.html', {
            'turma': turma,
            'aluno': aluno
        }
        )
        msg = EmailMultiAlternatives(subject, text_content, from_email, to)
        msg.attach_alternative(html_content, "text/html")
        msg.send()
    except Exception as E:
        print(E)
    else:
        print('email enviado com sucesso!')


@staff_member_required
def adm_cursos_cadastrar(request):
    form = CadastroCursoForm()

    if request.method == 'POST':
        form = CadastroCursoForm(request.POST, request.FILES)
        if form.is_valid():
            curso = form.save(commit=False)
            curso.user_inclusao = request.user
            curso.user_ultima_alteracao = request.user
            curso.save()

            messages.success(request, 'Novo curso cadastrado!')
            return redirect('cursos_empresariais:adm_cursos_listar')

    context = {
        'form': form,
        'CADASTRAR': 'NOVO'
    }
    return render(request, 'adm/cursos/adm_cursos_cad_edit.html', context)

@staff_member_required
def adm_curso_visualizar(request, id):
    curso = get_object_or_404(Curso, pk=id)
    turmas = Turma.objects.filter(curso=curso)
    context = {
        'curso': curso,
        'turmas': turmas,
        'CADASTRAR': 'NOVO'
    }
    return render(request, 'adm/cursos/adm_curso_visualizar.html', context)

@staff_member_required
def adm_curso_editar(request, id):
    curso = Curso.objects.get(id=id)
    form = CadastroCursoForm(instance=curso)

    if request.method == 'POST':
        form = CadastroCursoForm(request.POST, request.FILES, instance=curso)
        if form.is_valid():
            form.save()

    context = {
        'form': form,
        'CADASTRAR': 'EDITAR',
        'curso': curso
    }
    return render(request, 'adm/cursos/adm_cursos_cad_edit.html', context)

@staff_member_required
def adm_curso_detalhes(request, id):
    curso = Curso.objects.get(id=id)
    interessados = Alertar_Aluno_Sobre_Nova_Turma.objects.filter(curso=curso, alertado=False).order_by('aluno__pessoa__nome')
    matrizCur = Disciplinas.objects.filter(curso=curso)

    context = {
        'curso': curso,
        'interessados': interessados,
        'matrizesCur': matrizCur,
        'id': id
    }
    return render(request, 'adm/cursos/adm_cursos_detalhes.html', context)

@staff_member_required
def adm_cursos_interessados_excel(request, id):
    curso = Curso.objects.get(id=id)
    alunos = Alertar_Aluno_Sobre_Nova_Turma.objects.filter(curso=curso, alertado=False)
    response = HttpResponse(content_type='application/ms-excel')
    dt_now = timezone.now().strftime('%Y-%m-%d_%H-%M-%S')
    response['Content-Disposition'] = f'attachment; filename=interessados_{curso.nome}_{dt_now}.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = f'Alunos Interessados - {curso.nome}'
    ws.append(['Nome', 'Dt. inclusão', 'CPF', 'Email', 'Telefone', 'Bairro'])
    for aluno in alunos:
        ws.append([aluno.aluno.pessoa.nome, str(aluno.dt_inclusao), aluno.aluno.pessoa.cpf, aluno.aluno.pessoa.email, aluno.aluno.pessoa.telefone, aluno.aluno.pessoa.bairro])
    wb.save(response)
    return response



@staff_member_required
def cadastrar_categoria(request):

    form = CadastroCategoriaForm()
    if request.method == 'POST':
        form = CadastroCategoriaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nova categoria cadastrada!')
            return redirect('cursos_empresariais:adm_categorias_listar')
    context = {
        'form': form
    }
    return render(request, 'adm/cursos/cadastrar_categoria.html', context)

@staff_member_required
def remover_interessado(request, id):
    try:
        interessado = Alertar_Aluno_Sobre_Nova_Turma.objects.get(id=id)
        interessado.alertado = True
        interessado.save()
        return JsonResponse({'success': True})  # Retorno de sucesso como JSON
    except Alertar_Aluno_Sobre_Nova_Turma.DoesNotExist:
        return JsonResponse({'success': False})  # Retorno de falha como JSON

@staff_member_required
def cadastrar_local(request):
    form = CadastroLocalForm()

    if request.method == 'POST':
        form = CadastroLocalForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Novo local cadastrado!')
            return redirect('cursos_empresariais:adm_locais_listar')
    context = {
        'form': form
    }
    return render(request, 'adm/cursos/cadastrar_local.html', context)


@staff_member_required
def administrativo(request):
    return render(request, 'adm/administrativo.html')


@staff_member_required
def turmas(request):
    return render(request, 'adm/turmas/adm_turmas.html')


@staff_member_required
def adm_turmas_cadastrar(request):
    form = CadastroTurmaForm()

    if request.method == 'POST':
        form = CadastroTurmaForm(request.POST)
        if form.is_valid():
            turma = form.save(commit=False)
            turma.user_inclusao = request.user
            turma.user_ultima_alteracao = request.user
            turma.save()
            messages.success(request, 'Nova turma cadastrada com sucesso!')
            return redirect('cursos_empresariais:adm_turmas_listar')
    context = {
        'form': form
    }
    return render(request, 'adm/turmas/adm_turmas_cadastrar.html', context)


@staff_member_required
def adm_turmas_listar(request):

    turmas = Turma.objects.exclude(status='enc').order_by('data_final')

    context = {
        'turmas': turmas
    }
    return render(request, 'adm/turmas/adm_turmas_listar.html', context)

@staff_member_required
def adm_turmas_listar_encerradas(request):
    turmas = Turma.objects.filter(status="enc").order_by('-data_final')

    context = {
        'turmas': turmas
    }
    return render(request, 'adm/turmas/adm_turmas_listar_encerradas.html', context)


@staff_member_required
def adm_cursos_listar(request):
    cursos = Curso.objects.all()

    context = {
        'cursos': cursos
    }
    return render(request, 'adm/cursos/adm_cursos_listar.html', context)


@staff_member_required
def adm_locais(request):
    return render(request, 'adm/locais/adm_locais.html')


@staff_member_required
def adm_locais_cadastrar(request):
    form = CadastroLocalForm()
    if request.method == 'POST':
        form = CadastroLocalForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Novo local cadastrado!')
            return redirect('cursos_empresariais:adm_locais_listar')

    context = {
        'form': form,
        'CADASTRAR': 'NOVO'
    }
    return render(request, 'adm/locais/adm_locais_cadastrar.html', context)


@staff_member_required
def adm_locais_listar(request):
    locais = Local.objects.all()
    context = {
        'locais': locais
    }
    return render(request, 'adm/locais/adm_locais_listar.html', context)


@staff_member_required
def adm_locais_editar(request, id):
    local = Local.objects.get(id=id)
    form = CadastroLocalForm(instance=local)
    if request.method == 'POST':
        form = CadastroLocalForm(request.POST, instance=local)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Informações do local atualizada com sucesso')
            return redirect('cursos_empresariais:adm_locais_listar')

    context = {
        'form': form,
        'local': local
    }
    return render(request, 'adm/locais/adm_locais_editar.html', context)


@staff_member_required
def adm_locais_excluir(request, id):
    local = Local.objects.get(id=id)
    local.delete()
    return redirect('cursos_empresariais:adm_locais_listar')


@staff_member_required
def adm_categorias(request):
    return render(request, 'adm/categorias/adm_categorias.html')


@staff_member_required
def adm_categorias_cadastrar(request):
    form = CadastroCategoriaForm()
    if request.method == 'POST':
        form = CadastroCategoriaForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nova categoria cadastrada!')
            return redirect('cursos_empresariais:adm_categorias_listar')

    context = {
        'form': form,
        'CADASTRAR': 'NOVO'
    }
    return render(request, 'adm/categorias/adm_categorias_cadastrar.html', context)


@staff_member_required
def adm_categorias_listar(request):
    categorias = Categoria.objects.all()
    context = {
        'categorias': categorias
    }
    return render(request, 'adm/categorias/adm_categorias_listar.html', context)


@staff_member_required
def adm_categorias_excluir(request, id):
    categoria = Categoria.objects.get(id=id)
    categoria.delete()
    return redirect('cursos_empresariais:adm_categorias_listar')


@staff_member_required
def adm_categorias_editar(request, id):
    categoria = Categoria.objects.get(id=id)
    form = CadastroCategoriaForm(instance=categoria)
    if request.method == 'POST':
        form = CadastroCategoriaForm(request.POST, instance=categoria)
        if form.is_valid():
            form.save()
            messages.success(request, 'Informações da categoria atualizada!')
            return redirect('cursos_empresariais:adm_categorias_listar')

    context = {
        'form': form,
        'categoria': categoria
    }
    return render(request, 'adm/categorias/adm_categorias_editar.html', context)


@staff_member_required
def adm_instituicoes_listar(request):
    instituicoes = Instituicao.objects.all()
    context = {
        'instituicoes': instituicoes
    }
    return render(request, 'adm/instituicoes/adm_instituicoes_listar.html', context)


@staff_member_required
def adm_instituicao_cadastrar(request):
    form = Instituicao_form()
    if request.method == 'POST':
        form = Instituicao_form(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Nova instituição cadastrada!')
            return redirect('cursos_empresariais:adm_instituicoes_listar')

    context = {
        'form': form,
        'CADASTRAR': 'NOVO'
    }
    return render(request, 'adm/instituicoes/adm_instituicao_cadastrar.html', context)


@staff_member_required
def adm_turno_cadastrar(request, id):

    turma = get_object_or_404(Turma, pk=id)
    form = Turno_form()

    if request.method == 'POST':
        form = Turno_form(request.POST)
        if form.is_valid():
            turno = form.save()

            Turno_estabelecido.objects.create(turno=turno, turma=turma)

            messages.success(request, 'Novo turno cadastrado!')
            return redirect('cursos_empresariais:adm_turma_visualizar', turma.id)

    context = {
        'form': form,
        'CADASTRAR': 'NOVO'
    }
    return render(request, 'adm/turnos/adm_turno_cadastrar.html', context)


@staff_member_required
def adm_professores(request):
    context = {}
    return render(request, 'adm/professores/adm_professores.html', context)


@staff_member_required
def adm_professores_cadastrar(request):
    form = CadastroProfessorForm()
    if request.method == 'POST':
        form = CadastroProfessorForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Novo Instrutor cadastrada com sucesso!')
            return redirect('cursos_empresariais:adm_professores_listar')

    context = {
        'form': form,
    }
    return render(request, 'adm/professores/adm_professores_cadastrar.html', context)


@staff_member_required
def adm_professores_listar(request):
    instrutores = Instrutor.objects.all()
    context = {
        'Instrutores': instrutores
    }
    return render(request, 'adm/professores/adm_professores_listar.html', context)


@staff_member_required
def adm_professores_editar(request, id):
    instrutor = Instrutor.objects.get(id=id)
    form = CadastroProfessorForm(instance=instrutor)
    if request.method == 'POST':
        form = CadastroProfessorForm(request.POST, instance=instrutor)
        if form.is_valid():
            form.save()
            messages.success(
                request, 'Informações do Instrutor atualizadas com sucesso!')
            return redirect('cursos_empresariais:adm_professores_Listar')

    context = {
        'form': form,
        'instrutor': instrutor
    }
    return render(request, 'adm/professores/adm_professores_editar.html', context)


@staff_member_required
def adm_professores_excluir(request, id):
    instrutor = Instrutor.objects.get(id=id)
    instrutor.delete()
    return redirect('cursos_empresariais:adm_professores_listar')

@staff_member_required
def gerar_certificados(request, id):
    data_atual = datetime.date.today()
    turma = get_object_or_404(Turma, id=id)
    matriculas=Matricula.objects.filter(turma_id=id)
    disciplinas = Disciplinas.objects.filter(curso_id=turma.curso.id)
    aux=[0,0]
    for d in disciplinas:
        aux[0]+=int(d.n_aulas)
        aux[1]+=int(d.carga_horaria)
    context={
        'turma': turma,
        'matriculas': matriculas,
        'data_atual': data_atual,
        'instrutor': turma.instrutores.all()[0],
        'disciplinas': disciplinas,
        'total_aulas': aux[0],
        'total_horas': aux[1]
    }
    return render(request, 'certificados.html', context)

@staff_member_required
def adm_turmas_visualizar(request, id):
    turma = Turma.objects.get(id=id)

    matriculas = Matricula.objects.filter(turma=turma)

    matriculas_alunos = matriculas.filter(status='a').select_related('aluno')

    total_aulas = Aula.objects.filter(
        associacao_turma_turno__turma=turma).count()

    matriculas_alunos_array = []
    for matricula in matriculas_alunos:
        presencas = Presenca.objects.filter(matricula=matricula.matricula).count()
        frequencia = "Nenhuma aula registrada"
        if total_aulas:
            frequencia = f"{presencas/total_aulas * 100}%"

        matriculas_alunos_array.append(
            {'aluno': matricula.aluno, 'matricula': matricula, 'frequencia': frequencia})

    matriculas_selecionados = matriculas.filter(
        status='s').select_related('aluno')

    matriculas_candidatos = matriculas.filter(
        status='c').select_related('aluno')

    if request.method == 'POST':
        for i in request.POST.getlist("candidatos_selecionados"):
            if i != 'csrfmiddlewaretoken':
                matricula_candidato = Matricula.objects.get(pk=i)
                matricula_candidato.status = 's'
                matricula_candidato.save()

    context = {
        'total_aulas': total_aulas,
        'turma': turma,
        'matriculas_alunos': matriculas_alunos_array,
        'matriculas_selecionados': matriculas_selecionados,
        'matriculas_candidatos': matriculas_candidatos,
        'qnt_alunos': matriculas_alunos.count(),
        'qnt_alunos_espera': matriculas_candidatos.count() + matriculas_selecionados.count(),
        'is_cheio': turma.quantidade_permitido <= matriculas_alunos.count(),
        'realocar': turma.status == 'pre' and len(Matricula.objects.filter(turma__curso=turma.curso, status='r')) > 0,

    }

    return render(request, 'adm/turmas/adm_turma_visualizar.html', context)


@staff_member_required
def visualizar_turma_editar(request, id):
    
    turma = Turma.objects.get(id=id)
    form = CadastroTurmaForm(instance=turma)

    if request.method == 'POST':
        form = CadastroTurmaForm(request.POST, instance=turma)
        if form.is_valid():
            turma=form.save()
            if turma.status == 'ati':
                matriculas = Matricula.objects.filter(turma=turma)
                for matricula in matriculas:
                    if matricula.status == 's' or matricula.status == 'c':
                        matricula.status = 'r'
                        matricula.save()
            messages.success(request, 'Turma editada com sucesso!')
            return redirect('cursos_empresariais:adm_turma_visualizar', id)
        
    context = {
        'turma': turma,
        'form': form
    }
    return render(request, 'adm/turmas/adm_turmas_editar_turma.html', context)


@staff_member_required
def visualizar_turma_selecionado(request, matricula):
    matricula = Matricula.objects.get(pk=matricula)
    turma = Turma.objects.get(pk=matricula.turma_id)

    if turma.quantidade_permitido <= Matricula.objects.filter(turma=turma, status='a').count():
        messages.error(
            request, 'Turma cheia! Não é possível adicionar mais alunos.')
        return redirect('cursos_empresariais:adm_turma_visualizar', matricula.turma.id)

    birthDate = matricula.aluno.pessoa.dt_nascimento
    today = date.today()
    age = 99
    
    if birthDate:
        age = today.year - birthDate.year - \
            ((today.month, today.day) < (birthDate.month, birthDate.day))

    form = CadastroAlunoForm(instance=matricula.aluno, prefix='aluno')
    form_responsavel = ''

    if age < 18:
        responsavel = Responsavel.objects.get(aluno=matricula.aluno)
        form_responsavel = CadastroResponsavelForm(
            instance=responsavel, prefix='responsavel')

    if request.method == 'POST':
        form_aluno = CadastroAlunoForm(
            request.POST, instance=matricula.aluno, prefix='aluno')

        if form_aluno.is_valid():

            if form_responsavel != '':
                form_responsavel = CadastroResponsavelForm(
                    instance=responsavel, prefix='responsavel')
                if form_responsavel.is_valid():
                    form_responsavel.save()
                else:
                    raise Exception('Erro no form do responsável')

            aluno = form_aluno.save()
            matricula.status = 'a'
            matricula.save()

            messages.success(request, "Candidato selecionado cadastrado como aluno!")
        return redirect('cursos_empresariais:adm_turma_visualizar', matricula.turma.id)

    context = {
        'form': form,
        'form_responsavel': form_responsavel,
        'turma': turma,
        'selecionado': matricula.aluno,
        'matricula': matricula
    }
    return render(request, 'adm/turmas/adm_turmas_editar_selecionado.html', context)


@staff_member_required
def excluir_turma(request, id):
    turma = Turma.objects.get(id=id)

    turma.delete()

    return redirect('cursos_empresariais:adm_turmas_listar')

@staff_member_required
def matricular_aluno(request, id):
    if request.method == 'POST':
        form = MatriculaAlunoForm(request.POST)
        if form.is_valid():
            matricula = form.save(commit=False)
            matricula.dt_inclusao = datetime.datetime.now()
            matricula.dt_ultima_atualizacao = datetime.datetime.now()
            matricula.save()
            messages.success(request, 'Aluno matriculado com sucesso!')
            return redirect('cursos_empresariais:adm_aluno_visualizar', matricula.aluno.id)
    context={
        'form': MatriculaAlunoForm(initial={'aluno': id}),
    }
    return render(request, 'adm/alunos/adm_aluno_matricular.html', context)
@staff_member_required
def adm_realocar(request, id):
    turma = Turma.objects.get(id=id)
    if request.method == "POST":
        candidatos_selecionados = request.POST.getlist('candidatos_selecionados')
        for candidato in candidatos_selecionados:
            matricula_antiga = Matricula.objects.get(matricula=candidato)
            matricula_antiga.status = 'd'
            try:
                matricula_nova = Matricula.objects.create(aluno=matricula_antiga.aluno, turma=turma, status='c')
                matricula_nova.save()
                matricula_antiga.save()
                messages.success(request, f'Aluno(s) realocados para a turma {turma} com sucesso!')
            except IntegrityError as e:
                if str(e) == 'UNIQUE constraint failed: cursos_matricula.matricula':
                    messages.warning(request, f'Alguns alunos realocados <b>já estão matriculados</b> na turma {turma}!')
                    return redirect('cursos_empresariais:adm_turma_visualizar', turma.id)
        return redirect('cursos_empresariais:adm_turma_visualizar', turma.id)
    matriculas = Matricula.objects.filter(turma__curso=turma.curso, status='r').order_by('aluno__pessoa__nome')
            
    context={
        'turma': turma,
        'candidatos': matriculas
    }
    return render(request, 'adm/turmas/adm_turma_realocar.html', context)

@staff_member_required
def adm_alunos_listar(request):
    if request.method == 'POST':                
        alunos = Aluno.objects.filter(pessoa__nome__icontains=request.POST['pesquisa'])
        if alunos.count() == 0:
            alunos = Aluno.objects.filter(pessoa__cpf__icontains=request.POST['pesquisa'])
            if alunos.count() == 0:
                messages.warning(request, 'Nenhum aluno encontrado')
    else:
        alunos = Aluno.objects.all()
    paginator = Paginator(alunos, 35)
    context = {
        'alunos': paginator.get_page(request.GET.get('page')),
        'total_alunos': alunos.count(),
    }
    return render(request, 'adm/alunos/adm_alunos_listar.html', context)


@staff_member_required
def adm_aluno_visualizar(request, id):
    aluno = Aluno.objects.get(pk=id)
    responsavel = ''

    try:
        responsavel = Responsavel.objects.get(aluno=aluno)
    except:
        pass

    context = {
        'aluno': aluno,
        'matriculas': Matricula.objects.filter(aluno=aluno),
        'responsavel': responsavel,
    }

    return render(request, 'adm/alunos/adm_aluno_visualizar.html', context)


@staff_member_required
def adm_aluno_editar(request, id):
    aluno = Aluno.objects.get(pk=id)

    form = CadastroAlunoForm(instance=aluno)
    if request.method == 'POST':
        form = CadastroAlunoForm(request.POST, instance=aluno)
        if form.is_valid():
            form.save()
            messages.success(request, 'Aluno(a) editado(a) com sucesso!')
            return redirect('cursos_empresariais:adm_aluno_visualizar', id)

    context = {
        'aluno': aluno,
        'form': form
    }

    return render(request, 'adm/alunos/adm_aluno_editar.html', context)


@staff_member_required
def adm_aluno_excluir(request, id):
    aluno = Aluno.objects.get(id=id)

    aluno.delete()
    messages.success(request, 'Aluno excluido com sucesso')

    return redirect('cursos_empresariais:adm_alunos_listar')


@staff_member_required
def desmatricular_aluno(request, matricula):

    matricula_obj = Matricula.objects.get(matricula=matricula)
    matricula_obj.status = 'd'
    matricula_obj.save()
    messages.success(request, 'Aluno desmatriculado com sucesso')

    return redirect('cursos_empresariais:adm_aluno_visualizar', matricula_obj.aluno.id)


def calculate_age(born):
    today = date.today()
    return today.year - born.year - ((today.month, today.day) < (born.month, born.day))


@staff_member_required
def adm_aula_cadastrar(request, turma_id):

    turma = get_object_or_404(Turma, pk=turma_id)
    turno_choices = [(turno.id, turno)
                     for turno in Turno_estabelecido.objects.filter(turma=turma)]
    form = Aula_form()
    form.fields['associacao_turma_turno'].choices = turno_choices

    if request.method == 'POST':
        form = Aula_form(data=request.POST)
        form.fields['associacao_turma_turno'].choices = turno_choices

        if form.is_valid():
            aula = form.save()
            messages.success(request, 'Aula registra!')
            return redirect('cursos_empresariais:adm_aulas_listar', turma.id)

    context = {'form': form, 'CADASTRAR': 'NOVO'}
    return render(request, 'adm/aulas/adm_aula_cadastrar.html', context)


@staff_member_required
def adm_aulas_listar(request, turma_id):

    turma = get_object_or_404(Turma, pk=turma_id)
    aulas = Aula.objects.filter(associacao_turma_turno__turma=turma)

    context = {
        'turma': turma,
        'aulas': aulas
    }

    return render(request, 'adm/aulas/adm_aulas_listar.html', context)


@staff_member_required
def adm_aula_visualizar(request, turma_id, aula_id):

    if request.method == "POST":
        acao = request.POST.get('acao') or 'p'
        for matricula in request.POST.getlist('alunos_selecionados'):
            presenca = Presenca.objects.get_or_create(
                matricula=Matricula.objects.get(matricula=matricula), aula_id=aula_id)[0]
            presenca.status = acao
            presenca.save()

    turma = get_object_or_404(Turma, pk=turma_id)
    aula = get_object_or_404(Aula, pk=aula_id)
    matriculas = Matricula.objects.filter(turma=turma)

    matriculados = []
    for matricula in matriculas:
        try:
            presenca = Presenca.objects.get(aula=aula, matricula=matricula)
        except:
            presenca = ''

        matriculados.append({'matricula': matricula, 'presenca': presenca})

    context = {
        'turma': turma,
        'matriculados': matriculados,
        'aula': aula,
    }

    return render(request, 'adm/aulas/adm_aula_visualizar.html', context)


@staff_member_required
def adm_justificativa_cadastrar(request, presenca_id):

    form = Justificativa_form()
    presenca = get_object_or_404(Presenca, pk=presenca_id)

    if request.method == "POST":
        form = Justificativa_form(request.POST)
        if form.is_valid():
            justificativa = form.save()
            presenca.justificativa = justificativa
            presenca.save()

            messages.error(request, 'Justificativa registrada!')
            return redirect('cursos_empresariais:adm_aula_visualizar', presenca.aula.associacao_turma_turno.turma.id, presenca.aula.id)

    context = {
        'presenca': presenca,
        'form': form
    }

    return render(request, 'adm/justificativas/adm_justificativa_cadastrar.html', context)


@staff_member_required
def adm_justificativa_visualizar(request, presenca_id):

    presenca = get_object_or_404(Presenca, pk=presenca_id)

    context = {
        'presenca': presenca,
        'aluno': presenca.matricula.aluno
    }

    return render(request, 'adm/justificativas/adm_justificativa_visualizar.html', context)


@login_required
@user_passes_test(lambda u: u.is_superuser)
def adm_cadastro_aluno(request):
    form = Form_Alterar_Pessoa()
    if request.method == 'POST':
        try:
            pessoa=Pessoa.objects.get(cpf=request.POST['cpf'])
            # messages.error(request, 'Usuário já cadastrado')
        except:
            pessoa = None
            
        if not pessoa:
            form = Form_Alterar_Pessoa(request.POST)
            if form.is_valid():
                pessoa=form.save()
                partes=request.POST['dt_nascimento'].split('-')
                user = User.objects.create_user(username=str(validate_cpf(request.POST['cpf'])), first_name=request.POST['nome'] ,email=request.POST['email'] or None, password=partes[2] + partes[1] + partes[0])
                print(partes[2] + partes[1] + partes[0])
                pessoa.user=user
                pessoa.save()
                Aluno.objects.create(
                    pessoa=pessoa,
                    profissão='Não informado',
                    escolaridade='emc',
                    estado_civil='s',
                    aceita_mais_informacoes=True,
                    li_e_aceito_termos=True
                )
                # aluno.save()
                messages.success(request, 'Usuário e aluno cadastrado com sucesso!')
    else:
        try:
            aluno = Aluno.objects.get(pessoa=pessoa)
        except:
            aluno = None
        if not aluno:
            form = Form_Alterar_Pessoa(request.POST)
            if form.is_valid():
                Aluno.objects.create(
                            pessoa=pessoa,
                            profissão='Não informado',
                            escolaridade='emc',
                            estado_civil='s',
                            aceita_mais_informacoes=True,
                            li_e_aceito_termos=True
                        )
                messages.success(request, 'Pessoa cadastrada como aluno com sucesso!')
        else:
            messages.warning(request, 'Este aluno já está cadastrado')
        
    context = {
        'form': form
    }
    return render(request, 'adm/adm_cadastro.html', context) 